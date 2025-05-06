import os
from datetime import datetime, timedelta
from decimal import Decimal

import openpyxl
from django.conf import settings
from django.core.exceptions import SuspiciousOperation, ValidationError
from django.core.mail import EmailMessage
from django.db import transaction
from django.db.models import Avg, Case, Count, F, Prefetch, Q, Sum, When
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django_filters import rest_framework as filters
from django_q.tasks import async_task
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import filters as rf_filters
from rest_framework import mixins, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import APIException, AuthenticationFailed
from rest_framework.exceptions import ValidationError as RESTValidationError
from rest_framework.generics import get_object_or_404
from rest_framework.response import Response

from apps.aggregator.views import qs_to_xls
from apps.bank.models import BankAccount
from apps.bank.serializers import BankAccountSerializer
from apps.ledger.models import Account, Party
from apps.ledger.serializers import (
    AccountSerializer,
    JournalEntriesSerializer,
    PartyMinSerializer,
    SalesJournalEntrySerializer,
)
from apps.product.models import Item, Unit
from apps.product.serializers import (
    ItemPOSSerializer,
    ItemPurchaseSerializer,
    ItemSalesSerializer,
)
from apps.tax.models import TaxScheme
from apps.tax.serializers import TaxSchemeMinSerializer
from apps.users.serializers import CompanySerializer, FiscalYearSerializer
from apps.voucher.filters import (
    ChallanFilterSet,
    CreditNoteFilterSet,
    DebitNoteFilterSet,
    JournalVoucherFilterSet,
    PaymentReceiptFilterSet,
    PurchaseDiscountFilterSet,
    PurchaseOrderFilterSet,
    PurchaseVoucherFilterSet,
    PurchaseVoucherRowFilterSet,
    SalesDiscountFilterSet,
    SalesRowFilterSet,
    SalesVoucherFilterSet,
)
from apps.voucher.models import (
    Challan,
    ChallanRow,
    PaymentReceipt,
    SalesAgent,
)
from apps.voucher.resources import (
    CreditNoteResource,
    CreditNoteRowResource,
    DebitNoteResource,
    DebitNoteRowResource,
    PurchaseVoucherResource,
    PurchaseVoucherRowResource,
    SalesVoucherResource,
    SalesVoucherRowResource,
)
from apps.voucher.serializers.debit_note import (
    DebitNoteCreateSerializer,
    DebitNoteDetailSerializer,
    DebitNoteListSerializer,
)
from apps.voucher.serializers.purchase import (
    PurchaseOrderCreateSerializer,
    PurchaseOrderListSerializer,
    PurchaseVoucherRowSerializer,
)
from apps.voucher.serializers.sales import (
    EmailInvoiceRequestSerializer,
    ImportRequestSerializer,
    RecurringVoucherTemplateCreateSerializer,
)
from apps.voucher.serializers.voucher_settings import (
    PurchaseCreateSettingSerializer,
    PurchaseSettingCreateSerializer,
    PurchaseSettingSerializer,
    PurchaseUpdateSettingSerializer,
    SalesCreateSettingSerializer,
    SalesSettingCreateSerializer,
    SalesSettingsSerializer,
    SalesUpdateSettingSerializer,
)

# from awecount.libs.db import DistinctSum
from awecount.libs import get_next_voucher_no
from awecount.libs.CustomViewSet import (
    CollectionViewSet,
    CompanyViewSetMixin,
    CRULViewSet,
    GenericSerializer,
)
from awecount.libs.exception import UnprocessableException
from awecount.libs.helpers import (
    check_verification_hash,
    deserialize_request,
    get_origin,
    get_verification_hash,
    serialize_request,
    upload_file,
)
from awecount.libs.mixins import (
    CancelCreditOrDebitNoteMixin,
    CancelPurchaseVoucherMixin,
    DeleteRows,
    InputChoiceMixin,
)
from awecount.libs.nepdate import ad2bs, ad2bs_str

from ..models import (
    CreditNote,
    CreditNoteRow,
    DebitNote,
    DebitNoteRow,
    Import,
    LandedCostRow,
    PaymentMode,
    PurchaseDiscount,
    PurchaseOrder,
    PurchaseOrderRow,
    PurchaseVoucher,
    PurchaseVoucherRow,
    RecurringVoucherTemplate,
    SalesDiscount,
    SalesVoucher,
    SalesVoucherRow,
)
from ..models.invoice_design import InvoiceDesign
from ..models.journal_vouchers import JournalVoucher, JournalVoucherRow
from ..serializers import (
    ChallanCreateSerializer,
    ChallanListSerializer,
    CreditNoteCreateSerializer,
    CreditNoteDetailSerializer,
    CreditNoteListSerializer,
    InvoiceDesignSerializer,
    JournalVoucherCreateSerializer,
    JournalVoucherDetailSerializer,
    JournalVoucherListSerializer,
    PaymentReceiptDetailSerializer,
    PaymentReceiptFormSerializer,
    PaymentReceiptSerializer,
    PurchaseBookExportSerializer,
    PurchaseBookSerializer,
    PurchaseDiscountSerializer,
    PurchaseVoucherCreateSerializer,
    PurchaseVoucherDetailSerializer,
    PurchaseVoucherListSerializer,
    SalesAgentSerializer,
    SalesBookExportSerializer,
    SalesBookSerializer,
    SalesDiscountMinSerializer,
    SalesDiscountSerializer,
    SalesRowSerializer,
    SalesVoucherChoiceSerializer,
    SalesVoucherCreateSerializer,
    SalesVoucherDetailSerializer,
    SalesVoucherListSerializer,
)


def send_sales_voucher_import_completion_email(
    request, new_invoices=[], error_message=None
):
    if error_message:
        subject = "Error in Importing Sales Invoices"
        message = f"""
        <p>We encountered an error while importing your sales invoices:</p>
        <p><strong>Error:</strong> {error_message}</p>
        <p>Please review the file you attempted to upload and ensure that it follows the correct format. If the issue persists, feel free to reach out to our support team for assistance.</p>
        <p>Thank you for your patience.</p>
        """
    else:
        origin = get_origin()
        subject = "Sales Invoices Imported Successfully"
        message = f"""
        <p>Your sales invoices have been successfully imported.</p>
        <p><strong>{len(new_invoices)} invoice(s)</strong> have been imported and are now available in your system.</p>
        <p>You can view and manage these invoices in your <a href="{origin}/{request.company.slug}/sales/vouchers">Sales Voucher List</a>.</p>
        <p>Thank you for using our service. If you need any further assistance, don't hesitate to contact us.</p>
        """

    email = EmailMessage(
        subject=subject,
        body=message,
        to=[request.user.email],
        from_email=settings.DEFAULT_FROM_EMAIL,
    )
    email.content_subtype = "html"
    email.send()


def send_purchase_voucher_import_completion_email(
    request, new_invoices=[], error_message=None
):
    if error_message:
        subject = "Error in Importing Purchase Invoices"
        message = f"""
        <p>We encountered an error while importing your purchase invoices:</p>
        <p><strong>Error:</strong> {error_message}</p>
        <p>Please review the file you attempted to upload and ensure that it follows the correct format. If the issue persists, feel free to reach out to our support team for assistance.</p>
        <p>Thank you for your patience.</p>
        """
    else:
        origin = get_origin()
        subject = "Purchase Invoices Imported Successfully"
        message = f"""
        <p>Your purchase invoices have been successfully imported.</p>
        <p><strong>{len(new_invoices)} invoice(s)</strong> have been imported and are now available in your system.</p>
        <p>You can view and manage these invoices in your <a href="{origin}/purchase-voucher/list/">Purchase Voucher List</a>.</p>
        <p>Thank you for using our service. If you need any further assistance, don't hesitate to contact us.</p>
        """

    email = EmailMessage(
        subject=subject,
        body=message,
        to=[request.user.email],
        from_email=settings.DEFAULT_FROM_EMAIL,
    )
    email.content_subtype = "html"
    email.send()


def import_sales_vouchers(request_obj, file):
    request = deserialize_request(request_obj)
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))

    required_fields = [
        "Invoice Group ID",
        "Party",
        "Customer Name",
        "Address",
        "Due Date",
        "Discount Type",
        "Discount",
        "Trade Discount",
        "Payment Mode",
        "Remarks",
        "Is Export",
        "Sales Agent ID",
        "Status",
        "Row Item ID",
        "Row Quantity",
        "Row Rate",
        "Row Unit ID",
        "Row Discount Type",
        "Row Discount",
        "Row Tax Scheme ID",
        "Row Description",
    ]
    headers = rows[0]
    error_message = None
    for i in range(len(required_fields)):
        if required_fields[i] != headers[i]:
            error_message = f"Column {i + 1} should be {required_fields[i]}"
            break

    if error_message:
        send_sales_voucher_import_completion_email(request, error_message=error_message)
        return

    data = rows[1:]
    invoices = {}

    for row in data:
        invoice_id = row[0]
        if invoice_id not in invoices:
            invoices[invoice_id] = {
                "date": timezone.now().date(),
                "party": row[1],
                "customer_name": row[2],
                "address": row[3],
                "due_date": row[4].date() if isinstance(row[4], datetime) else row[4],
                "discount_type": row[5],
                "discount": row[6] or 0,
                "trade_discount": row[7],
                "payment_mode": row[8],
                "remarks": row[9],
                "is_export": row[10],
                "sales_agent": row[11],
                "rows": [],
                "status": row[12],
            }
        invoices[invoice_id]["rows"].append(
            {
                "item_id": row[13],
                "quantity": row[14],
                "rate": row[15],
                "unit_id": row[16],
                "discount_type": row[17],
                "discount": row[18],
                "tax_scheme_id": row[19],
                "description": row[20],
            }
        )

    new_invoices = invoices.items()

    try:
        with transaction.atomic():
            for invoice_id, invoice_data in new_invoices:
                serializer = SalesVoucherCreateSerializer(
                    data=invoice_data, context={"request": request}
                )
                serializer.is_valid(raise_exception=True)
                instance = serializer.create(validated_data=serializer.validated_data)
                invoice_data["id"] = instance.id
    except RESTValidationError as e:
        error_message = e.detail
    except SuspiciousOperation as e:
        error_message = str(e)

    send_sales_voucher_import_completion_email(
        request, new_invoices=new_invoices, error_message=error_message
    )

    Import.objects.filter(
        company=request.company, type="Sales Voucher", status="Pending"
    ).update(status="Failed" if error_message else "Completed")


def import_purchase_vouchers(request_obj, file):
    request = deserialize_request(request_obj)
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))

    required_fields = [
        "Invoice Group ID",
        "Party",
        "Bill Number",
        "Discount Type",
        "Discount",
        "Trade Discount",
        "Date",
        "Payment Mode",
        "Due Date",
        "Remarks",
        "Is Import",
        "Status",
        "Row Item ID",
        "Row Quantity",
        "Row Rate",
        "Row Unit ID",
        "Row Discount Type",
        "Row Discount",
        "Row Tax Scheme ID",
        "Row Description",
    ]
    headers = rows[0]
    error_message = None
    for i in range(len(required_fields)):
        if required_fields[i] != headers[i]:
            error_message = f"Column {i + 1} should be {required_fields[i]}"
            break

    if error_message:
        send_purchase_voucher_import_completion_email(
            request, error_message=error_message
        )
        return

    data = rows[1:]
    invoices = {}

    for row in data:
        invoice_id = row[0]
        if invoice_id not in invoices:
            invoices[invoice_id] = {
                "party": row[1],
                "voucher_no": row[2],
                "discount_type": row[3],
                "discount": row[4] or 0,
                "trade_discount": row[5],
                "date": row[6].date() if isinstance(row[6], datetime) else row[6],
                "payment_mode": row[7],
                "due_date": row[8].date() if isinstance(row[8], datetime) else row[8],
                "remarks": row[9],
                "is_import": row[10],
                "status": row[11],
                "rows": [],
            }
        invoices[invoice_id]["rows"].append(
            {
                "item_id": row[12],
                "quantity": row[13],
                "rate": row[14],
                "unit_id": row[15],
                "discount_type": row[16],
                "discount": row[17],
                "tax_scheme_id": row[18],
                "description": row[19],
            }
        )

    new_invoices = invoices.items()

    try:
        with transaction.atomic():
            for invoice_id, invoice_data in new_invoices:
                serializer = PurchaseVoucherCreateSerializer(
                    data=invoice_data, context={"request": request}
                )
                serializer.is_valid(raise_exception=True)
                instance = serializer.create(validated_data=serializer.validated_data)
                invoice_data["id"] = instance.id

    except RESTValidationError as e:
        error_message = e.detail
    except SuspiciousOperation as e:
        error_message = str(e)

    send_purchase_voucher_import_completion_email(
        request, new_invoices=new_invoices, error_message=error_message
    )

    Import.objects.filter(
        company=request.company, type="Purchase Voucher", status="Pending"
    ).update(status="Failed" if error_message else "Completed")


class SalesVoucherViewSet(InputChoiceMixin, DeleteRows, CRULViewSet):
    queryset = SalesVoucher.objects.all()
    serializer_class = SalesVoucherCreateSerializer
    model = SalesVoucher
    row = SalesVoucherRow
    collections = [
        ("parties", Party, PartyMinSerializer, True, ["name"]),
        ("units", Unit, GenericSerializer, True, ["name"]),
        ("discounts", SalesDiscount, SalesDiscountMinSerializer, False),
        (
            "bank_accounts",
            BankAccount,
            GenericSerializer,
            True,
            ["bank_name", "short_name", "account_number"],
        ),
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_sales=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        ("tax_schemes", TaxScheme, TaxSchemeMinSerializer, True, ["name"]),
        (
            "items",
            Item.objects.filter(
                Q(can_be_sold=True) | Q(direct_expense=True)
            ).select_related("unit"),
            ItemSalesSerializer,
            True,
            ["name"],
        ),
    ]

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]

    filterset_class = SalesVoucherFilterSet

    search_fields = [
        "voucher_no",
        "party__name",
        "remarks",
        "total_amount",
        "party__tax_identification_number",
        "customer_name",
        "date",
        "address",
        "rows__item__name",
        "payment_mode__name",
        "bank_account__bank_name",
        "bank_account__short_name",
    ]

    def get_collections(self, request=None, *args, **kwargs):
        sales_agent_tuple = ("sales_agents", SalesAgent)
        if (
            request.company.enable_sales_agents
            and sales_agent_tuple not in self.collections
        ):
            # noinspection PyTypeChecker
            self.collections.append(sales_agent_tuple)
        return super().get_collections(request)

    def get_queryset(self, **kwargs):
        qs = super(SalesVoucherViewSet, self).get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows", "rows__item", "rows__unit")
        elif self.action == "list":
            qs = qs.select_related("party").prefetch_related(
                Prefetch(
                    "payment_receipts",
                    PaymentReceipt.objects.exclude(status="Cancelled"),
                    to_attr="receipts",
                )
            )
        return qs.order_by("-date", "-voucher_no")

    def get_serializer_class(self):
        if self.action == "choices":
            return SalesVoucherChoiceSerializer
        if self.action == "list":
            return SalesVoucherListSerializer
        return SalesVoucherCreateSerializer

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.is_issued():
            if not request.company.enable_sales_invoice_update:
                raise APIException({"detail": "Issued sales invoices can't be updated"})
            # self.request.user.check_perm("SalesIssuedModify")
        return super().update(request, *args, **kwargs)

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        sale_voucher = get_object_or_404(SalesVoucher, pk=pk)
        journals = sale_voucher.journal_entries()
        return Response(SalesJournalEntrySerializer(journals, many=True).data)

    def get_voucher_details(self, pk):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    SalesVoucherRow.objects.all()
                    .select_related(
                        "item", "item__category", "unit", "discount_obj", "tax_scheme"
                    )
                    .order_by("pk"),
                )
            )
            .prefetch_related(
                Prefetch(
                    "payment_receipts",
                    PaymentReceipt.objects.exclude(status="Cancelled"),
                    to_attr="receipts",
                )
            )
            .select_related(
                "discount_obj", "bank_account", "company__sales_setting", "party"
            )
        )
        data = SalesVoucherDetailSerializer(
            get_object_or_404(pk=pk, queryset=qs), context={"request": self.request}
        ).data
        data["can_update_issued"] = self.request.company.enable_sales_invoice_update
        data["available_payment_modes"] = GenericSerializer(
            PaymentMode.objects.filter(
                company=self.request.company, enabled_for_sales=True
            ),
            many=True,
        ).data
        return data

    @action(detail=True)
    def details(self, request, pk, *args, **kwargs):
        details = self.get_voucher_details(pk)
        hash = get_verification_hash("sales-invoice-{}".format(pk))
        return Response(
            {
                **details,
                "hash": hash,
            }
        )

    @action(detail=True, permission_classes=[], url_path="details-by-hash")
    def details_by_hash(self, request, pk):
        hash = request.GET.get("hash")
        if not hash:
            raise AuthenticationFailed("No hash provided")
        if check_verification_hash(hash, "sales-invoice-{}".format(pk)) is False:
            raise AuthenticationFailed("Invalid hash")
        obj = SalesVoucher.objects.get(pk=pk)
        self.request.company = obj.company
        self.request.company_id = obj.company_id
        details = self.get_voucher_details(pk)
        return Response({**details, "company": CompanySerializer(obj.company).data})

    @action(detail=True, url_path="email-invoice", methods=["POST"])
    def email_invoice(self, request, pk):
        serializer = EmailInvoiceRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = self.get_object()
        async_task(
            "apps.voucher.models.SalesVoucher.email_invoice",
            obj,
            **serializer.validated_data,
        )
        return Response({})

    def get_defaults(self, request=None, *args, **kwargs):
        return {
            "options": {
                "fiscal_years": FiscalYearSerializer(
                    request.company.get_fiscal_years(), many=True
                ).data,
                "enable_sales_agents": request.company.enable_sales_agents,
                "enable_fifo": request.company.inventory_setting.enable_fifo,
                "default_mode_obj": None
                if request.company.sales_setting.mode in ["Cash", "Credit"]
                else BankAccount.objects.filter(id=request.company.sales_setting.mode)
                .annotate(name=F("short_name") or F("bank_name") or F("account_number"))
                .values("id", "name")
                .first(),
            },
        }

    def get_create_defaults(self, request=None, *args, **kwargs):
        data = SalesCreateSettingSerializer(request.company.sales_setting).data
        data["options"]["voucher_no"] = get_next_voucher_no(
            SalesVoucher, request.company.id
        )
        return data

    def get_update_defaults(self, request=None, *args, **kwargs):
        data = SalesUpdateSettingSerializer(request.company.sales_setting).data
        data["options"]["can_update_issued"] = (
            request.company.enable_sales_invoice_update
        )
        obj = self.get_object()
        if not obj.voucher_no:
            data["options"]["voucher_no"] = get_next_voucher_no(
                SalesVoucher, request.company.id
            )
        return data

    @action(detail=True, methods=["POST"])
    def mark_as_paid(self, request, pk, *args, **kwargs):
        sale_voucher = self.get_object()
        try:
            sale_voucher.mark_as_resolved(status="Paid")
            return Response({})
        except Exception as e:
            raise APIException(str(e))

    @action(detail=True, methods=["POST"], url_path="update-mode")
    def update_mode(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        mode = request.data.get("mode")
        if mode and str(mode).isdigit():
            obj.bank_account_id = mode
            obj.mode = "Bank Deposit"
        else:
            obj.mode = mode
            obj.bank_account_id = None
        obj.apply_transactions()
        return Response({})

    @action(detail=True, methods=["POST"], url_path="update-payment-mode")
    def update_payment_mode(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        payment_mode = request.data.get("payment_mode")
        if payment_mode is not None:
            try:
                PaymentMode.objects.get(id=payment_mode, company=request.company)
            except PaymentMode.DoesNotExist:
                raise RESTValidationError({"payment_mode": "Invalid payment mode"})

        obj.payment_mode_id = payment_mode

        if payment_mode is None:
            obj.status = "Issued"

        obj.apply_transactions()
        return Response({})

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk, *args, **kwargs):
        sales_voucher = self.get_object()
        message = request.data.get("message")
        if not message:
            raise RESTValidationError(
                {"message": "message field is required for cancelling invoice!"}
            )

        if sales_voucher.credit_notes.exclude(status="Cancelled").exists():
            raise RESTValidationError(
                {
                    "message": "This sales voucher has credit notes. Please cancel them first."
                }
            )

        # FIFO inconsistency check
        if (
            request.company.inventory_setting.enable_fifo
            and not request.query_params.get("fifo_inconsistency")
        ):
            raise UnprocessableException(
                detail="This may cause inconsistencies in fifo!",
                code="fifo_inconsistency",
            )

        sales_voucher.cancel(message)
        return Response({})

    @action(detail=True, methods=["POST"], url_path="log-print")
    def log_print(self, request, pk, *args, **kwargs):
        sale_voucher = self.get_object()
        sale_voucher.print_count += 1
        sale_voucher.save()
        return Response({"print_count": sale_voucher.print_count})

    @action(detail=False, url_path="by-voucher-no")
    def by_voucher_no(self, request, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    SalesVoucherRow.objects.all().select_related(
                        "item", "item__category", "unit", "discount_obj", "tax_scheme"
                    ),
                )
            )
            .select_related("discount_obj", "bank_account")
        )
        return Response(
            SalesVoucherDetailSerializer(
                get_object_or_404(
                    voucher_no=request.query_params.get("invoice_no"),
                    fiscal_year_id=request.query_params.get("fiscal_year"),
                    queryset=qs,
                ),
                context={"request": self.request},
            ).data
        )

    @action(detail=False)
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        params = [
            ("Invoices", queryset, SalesVoucherResource),
            (
                "Sales Rows",
                SalesVoucherRow.objects.filter(voucher__company_id=request.company.id),
                SalesVoucherRowResource,
            ),
        ]
        return qs_to_xls(params)

    @action(detail=False, url_path="import", methods=["POST"])
    def import_xls(self, request):
        serializer = ImportRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        xls_file = serializer.validated_data.get("file")

        if Import.objects.filter(
            company=request.company, type="Sales Voucher", status="Pending"
        ).exists():
            raise RESTValidationError({"error": "There is already an import pending."})

        Import.objects.create(
            company=request.company,
            type="Sales Voucher",
            status="Pending",
            file=xls_file,
            user=request.user,
        )

        async_task(
            "apps.voucher.api.import_sales_vouchers",
            serialize_request(request),
            xls_file,
        )

        return Response({})


class RecurringVoucherTemplateViewSet(CRULViewSet):
    queryset = RecurringVoucherTemplate.objects.all()
    serializer_class = RecurringVoucherTemplateCreateSerializer
    model = RecurringVoucherTemplate

    filter_backends = [
        rf_filters.SearchFilter,
    ]

    search_fields = ["title"]

    collections = [
        ("parties", Party, PartyMinSerializer, True, ["name"]),
        ("units", Unit, GenericSerializer, True, ["name"]),
        (
            "bank_accounts",
            BankAccount,
            GenericSerializer,
            True,
            ["bank_name", "short_name", "account_number"],
        ),
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_sales=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        ("tax_schemes", TaxScheme, TaxSchemeMinSerializer, False),
        (
            "items",
            Item.objects.filter(
                Q(can_be_sold=True) | Q(direct_expense=True)
            ).select_related("unit"),
            ItemSalesSerializer,
            True,
            ["name"],
        ),
    ]

    def get_collections(self, request):
        type = request.GET.get("type")
        if type:
            if type == "Sales Voucher":
                self.collections.append(
                    ("discounts", SalesDiscount, SalesDiscountMinSerializer, False)
                )
            elif type == "Purchase Voucher":
                self.collections.append(
                    ("discounts", PurchaseDiscount, PurchaseDiscountSerializer, False)
                )
        return super().get_collections(request)

    def retrieve(self, request, *args, **kwargs):
        res = self.get_serializer(self.get_object()).data

        invoice_data = res.get("invoice_data")

        if invoice_data.get("party"):
            invoice_data["selected_party_obj"] = PartyMinSerializer(
                Party.objects.get(id=invoice_data.get("party"))
            ).data
        if invoice_data.get("sales_agent"):
            invoice_data["sales_agent"] = SalesAgentSerializer(
                SalesAgent.objects.get(id=invoice_data.get("sales_agent"))
            ).data

        rows = invoice_data.get("rows")

        if rows:
            item_ids = [row.get("item_id") for row in rows]
            unit_ids = [row.get("unit_id") for row in rows]

            items = {
                str(item.id): item for item in Item.objects.filter(id__in=item_ids)
            }
            units = {
                str(unit.id): unit for unit in Unit.objects.filter(id__in=unit_ids)
            }

            for row in rows:
                row["selected_item_obj"] = ItemSalesSerializer(
                    items[str(row.get("item_id"))]
                ).data
                row["selected_unit_obj"] = GenericSerializer(
                    units[str(row.get("unit_id"))]
                ).data

        return Response(res)

    def get_defaults(self, request=None):
        type = self.request.GET.get("type")
        if type == "Sales Voucher":
            data = SalesCreateSettingSerializer(request.company.sales_setting).data
        elif type == "Purchase Voucher":
            data = PurchaseCreateSettingSerializer(
                request.company.purchase_setting
            ).data
        else:
            data = {}
        return data

    def get_queryset(self, company_id=None):
        qs = super().get_queryset()
        type = self.request.GET.get("type")
        if type:
            qs = qs.filter(type=type)
        return qs.order_by("-created_at")


class POSViewSet(
    DeleteRows,
    CompanyViewSetMixin,
    CollectionViewSet,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    viewsets.GenericViewSet,
):
    queryset = SalesVoucher.objects.all()
    serializer_class = SalesVoucherCreateSerializer
    model = SalesVoucher
    collections = [
        (
            "units",
            Unit.objects.only("name", "short_name"),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "discounts",
            SalesDiscount.objects.only("name", "type", "value"),
            SalesDiscountMinSerializer,
            False,
        ),
        (
            "bank_accounts",
            BankAccount.objects.only("short_name", "bank_name", "account_number"),
            GenericSerializer,
            True,
            ["short_name", "account_number", "bank_name"],
        ),
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_sales=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "tax_schemes",
            TaxScheme.objects.only("name", "short_name", "rate"),
            TaxSchemeMinSerializer,
            False,
        ),
    ]

    def get_item_queryset(self, company_id):
        return (
            Item.objects.filter(can_be_sold=True, company_id=company_id)
            .annotate(
                sold_qty=Coalesce(
                    Sum(
                        "sales_rows__quantity",
                        filter=Q(
                            sales_rows__voucher__date__gte=timezone.now()
                            - timedelta(days=30)
                        ),
                    ),
                    Decimal("0"),
                )
            )
            .only("name", "unit_id", "selling_price", "tax_scheme_id", "code")
            .order_by("-sold_qty")
        )

    def get_collections(self, request=None, *args, **kwargs):
        data = super().get_collections(request)
        qs = self.get_item_queryset(request.company.id)
        self.paginator.page_size = settings.POS_ITEMS_SIZE
        page = self.paginate_queryset(qs)
        serializer = ItemPOSSerializer(page, many=True)
        data["items"] = self.paginator.get_response_data(serializer.data)

        # qs = qs[: self.POS_ITEMS_SIZE]
        # data['items'] = {'results': ItemPOSSerializer(qs, many=True).data, 'pagination': {'page': 1, 'size': settings.POS_ITEMS_SIZE}}
        return data

    def get_create_defaults(self, request=None, *args, **kwargs):
        data = SalesCreateSettingSerializer(request.company.sales_setting).data
        return data

    def perform_create(self, serializer):
        serializer.validated_data["company_id"] = self.request.company.id
        if serializer.validated_data["status"] != "Draft":
            serializer.validated_data["print_count"] = 1
        try:
            serializer.save()
        except ValidationError as e:
            raise APIException({"detail": e.messages})

    def update(self, request, *args, **kwargs):
        if self.get_object().is_issued():
            raise APIException({"detail": "Issued POS invoices can't be updated"})
        return super().update(request, *args, **kwargs)


class PurchaseVoucherViewSet(
    CancelPurchaseVoucherMixin, InputChoiceMixin, DeleteRows, CRULViewSet
):
    serializer_class = PurchaseVoucherCreateSerializer
    model = PurchaseVoucher
    row = PurchaseVoucherRow

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "voucher_no",
        "party__name",
        "remarks",
        "total_amount",
        "party__tax_identification_number",
        "rows__item__name",
    ]
    filterset_class = PurchaseVoucherFilterSet

    collections = (
        ("parties", Party, PartyMinSerializer, True, ["name"]),
        ("discounts", PurchaseDiscount, PurchaseDiscountSerializer, False),
        ("units", Unit, GenericSerializer, True, ["name"]),
        (
            "bank_accounts",
            BankAccount,
            GenericSerializer,
            True,
            ["short_name", "account_number", "bank_name"],
        ),
        ("tax_schemes", TaxScheme, TaxSchemeMinSerializer, True, ["name"]),
        (
            "bank_accounts",
            BankAccount,
            GenericSerializer,
            True,
            ["short_name", "account_number", "bank_name"],
        ),
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_purchase=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "items",
            Item.objects.filter(
                Q(can_be_purchased=True)
                | Q(direct_expense=True)
                | Q(indirect_expense=True)
                | Q(fixed_asset=True)
            ).select_related("unit"),
            ItemPurchaseSerializer,
            True,
            ["name"],
        ),
    )

    def create(self, request, *args, **kwargs):
        voucher_no = request.data.get("voucher_no", None)
        party_id = request.data.get("party", None)
        fiscal_year = request.company.current_fiscal_year
        if self.model.objects.filter(
            voucher_no=voucher_no, party_id=party_id, fiscal_year=fiscal_year
        ).exists():
            raise ValidationError(
                {
                    "voucher_no": [
                        "Purchase with the bill number for the chosen party already exists."
                    ]
                }
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        voucher_no = request.data.get("voucher_no", None)
        party_id = request.data.get("party", None)
        fiscal_year = request.company.current_fiscal_year
        if (
            self.model.objects.filter(
                voucher_no=voucher_no, party_id=party_id, fiscal_year=fiscal_year
            )
            .exclude(id=obj.id)
            .exists()
        ):
            raise ValidationError(
                {
                    "voucher_no": [
                        "Purchase with the bill number for the chosen party already exists."
                    ]
                }
            )
        return super().update(request, *args, **kwargs)

    def get_create_defaults(self, request=None, *args, **kwargs):
        return PurchaseCreateSettingSerializer(request.company.purchase_setting).data

    def get_update_defaults(self, request=None, *args, **kwargs):
        data = PurchaseUpdateSettingSerializer(request.company.purchase_setting).data
        obj = self.get_object()
        if not obj.voucher_no:
            data["options"]["voucher_no"] = get_next_voucher_no(
                PurchaseVoucher, request.company.id
            )
        return data

    def get_queryset(self, **kwargs):
        qs = super().get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows", "rows__item", "rows__unit")
        elif self.action == "list":
            qs = qs.select_related("party")
        return qs.order_by("-date", "-pk")

    def get_serializer_class(self):
        if self.action == "list" or self.action in ("choices",):
            return PurchaseVoucherListSerializer
        return PurchaseVoucherCreateSerializer

    def get_defaults(self, request=None, *args, **kwargs):
        return {
            "options": {
                "fiscal_years": FiscalYearSerializer(
                    request.company.get_fiscal_years(), many=True
                ).data,
                "default_mode_obj": None
                if request.company.purchase_setting.mode in ["Cash", "Credit"]
                else BankAccount.objects.filter(
                    id=request.company.purchase_setting.mode
                )
                .annotate(name=F("short_name") or F("bank_name") or F("account_number"))
                .values("id", "name")
                .first(),
                "default_payment_mode_obj": None,
            },
        }

    @action(detail=True)
    def details(self, request, pk, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    PurchaseVoucherRow.objects.all()
                    .select_related(
                        "item", "item__category", "unit", "discount_obj", "tax_scheme"
                    )
                    .order_by("pk"),
                ),
                Prefetch(
                    "landed_cost_rows",
                    LandedCostRow.objects.all().order_by("pk"),
                ),
            )
            .select_related(
                "discount_obj", "bank_account", "company__purchase_setting", "party"
            )
        )
        return Response(
            PurchaseVoucherDetailSerializer(get_object_or_404(pk=pk, queryset=qs)).data
        )

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        purchase_voucher = get_object_or_404(PurchaseVoucher, pk=pk)
        journals = purchase_voucher.journal_entries()
        return Response(SalesJournalEntrySerializer(journals, many=True).data)

    @action(detail=False, url_path="by-voucher-no")
    def by_voucher_no(self, request, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    PurchaseVoucherRow.objects.all().select_related(
                        "item", "unit", "discount_obj", "tax_scheme"
                    ),
                )
            )
            .select_related("discount_obj", "bank_account")
        )
        voucher = get_object_or_404(
            voucher_no=request.query_params.get("invoice_no"),
            party_id=request.query_params.get("party"),
            fiscal_year_id=request.query_params.get("fiscal_year"),
            queryset=qs,
        )

        return Response(PurchaseVoucherDetailSerializer(voucher).data)

    @action(detail=True, methods=["POST"])
    def mark_as_paid(self, request, pk, *args, **kwargs):
        purchase_voucher = self.get_object()
        try:
            purchase_voucher.mark_as_resolved(status="Paid")
            return Response({})
        except Exception as e:
            raise APIException(str(e))

    @action(detail=False)
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        params = [
            ("Invoices", queryset, PurchaseVoucherResource),
            (
                "Purchase Rows",
                PurchaseVoucherRow.objects.filter(
                    voucher__company_id=request.company.id
                ),
                PurchaseVoucherRowResource,
            ),
        ]
        return qs_to_xls(params)

    @action(detail=False, url_path="import", methods=["POST"])
    def import_xls(self, request):
        serializer = ImportRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        xls_file = serializer.validated_data.get("file")

        if Import.objects.filter(
            company=request.company, type="Purchase Voucher", status="Pending"
        ).exists():
            raise RESTValidationError({"error": "There is already an import pending."})

        Import.objects.create(
            company=request.company,
            type="Purchase Voucher",
            status="Pending",
            file=xls_file,
            user=request.user,
        )

        async_task(
            "apps.voucher.api.import_purchase_vouchers",
            serialize_request(request),
            xls_file,
        )

        return Response({})


class CreditNoteViewSet(DeleteRows, CRULViewSet, CancelCreditOrDebitNoteMixin):
    serializer_class = CreditNoteCreateSerializer
    model = CreditNote
    row = CreditNoteRow

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "voucher_no",
        "party__name",
        "remarks",
        "total_amount",
        "party__tax_identification_number",
        "rows__item__name",
    ]
    filterset_class = CreditNoteFilterSet

    collections = (
        ("discounts", SalesDiscount, SalesDiscountSerializer, False),
        ("units", Unit, GenericSerializer, True, ["name"]),
        (
            "bank_accounts",
            BankAccount,
            GenericSerializer,
            True,
            ["short_name", "account_number", "bank_name"],
        ),
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_purchase=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        ("tax_schemes", TaxScheme, TaxSchemeMinSerializer, True, ["name"]),
        (
            "bank_accounts",
            BankAccount,
            BankAccountSerializer,
            True,
            ["short_name", "account_number", "bank_name"],
        ),
        (
            "items",
            Item.objects.filter(can_be_sold=True).select_related("unit"),
            ItemSalesSerializer,
            True,
            ["name"],
        ),
    )

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows", "rows__item", "rows__unit")
        elif self.action == "list":
            qs = qs.select_related("party")
        return qs.order_by("-id")

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.is_issued():
            if not request.company.enable_credit_note_update:
                raise APIException({"detail": "Issued credit notes can't be updated"})
            # self.request.user.check_perm("CreditNoteIssuedModify")
        return super().update(request, *args, **kwargs)

    def get_serializer_class(self):
        if self.action == "list":
            return CreditNoteListSerializer
        if self.action == "retrieve":
            return CreditNoteDetailSerializer
        return CreditNoteCreateSerializer

    def get_defaults(self, request=None, *args, **kwargs):
        data = {
            "options": {
                "fiscal_years": FiscalYearSerializer(
                    request.company.get_fiscal_years(), many=True
                ).data
            },
            "fields": {"can_update_issued": request.company.enable_credit_note_update},
        }
        return data

    def get_create_defaults(self, request=None, *args, **kwargs):
        options = SalesCreateSettingSerializer(request.company.sales_setting).data
        options["voucher_no"] = get_next_voucher_no(CreditNote, request.company.id)
        return {"options": options}

    def get_update_defaults(self, request=None, *args, **kwargs):
        options = SalesUpdateSettingSerializer(request.company.sales_setting).data

        obj = self.get_object()
        invoice_objs = []
        for inv in obj.invoices.all():
            invoice_objs.append({"id": inv.id, "voucher_no": inv.voucher_no})
        options["sales_invoice_objs"] = invoice_objs
        if not obj.voucher_no:
            options["voucher_no"] = get_next_voucher_no(
                SalesVoucher, request.company.id
            )

        return {"options": options}

    @action(detail=True)
    def details(self, request, pk, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                "invoices",
                Prefetch(
                    "rows",
                    CreditNoteRow.objects.all()
                    .select_related(
                        "item", "item__category", "unit", "discount_obj", "tax_scheme"
                    )
                    .order_by("pk"),
                ),
            )
            .select_related("discount_obj", "bank_account", "party")
        )
        data = CreditNoteDetailSerializer(get_object_or_404(pk=pk, queryset=qs)).data
        data["can_update_issued"] = request.company.enable_credit_note_update
        return Response(data)

    @action(detail=True, methods=["POST"])
    def mark_as_resolved(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        try:
            obj.mark_as_resolved()
            return Response({})
        except Exception as e:
            raise APIException(str(e))

    @action(detail=True, methods=["POST"], url_path="log-print")
    def log_print(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        obj.print_count += 1
        obj.save()
        return Response({"print_count": obj.print_count})

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        credit_note = get_object_or_404(CreditNote, pk=pk)
        journals = credit_note.journal_entries()
        return Response(JournalEntriesSerializer(journals, many=True).data)

    @action(detail=False)
    def export(self, request, *args, **kwargs):
        params = [
            ("Invoices", self.get_queryset(), CreditNoteResource),
            (
                "Credit Note Rows",
                CreditNoteRow.objects.filter(voucher__company_id=request.company.id),
                CreditNoteRowResource,
            ),
        ]
        return qs_to_xls(params)


class DebitNoteViewSet(DeleteRows, CRULViewSet, CancelCreditOrDebitNoteMixin):
    serializer_class = DebitNoteCreateSerializer
    model = DebitNote
    row = DebitNoteRow

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "voucher_no",
        "party__name",
        "remarks",
        "total_amount",
        "party__tax_identification_number",
        "rows__item__name",
    ]
    filterset_class = DebitNoteFilterSet

    collections = (
        ("discounts", PurchaseDiscount, PurchaseDiscountSerializer, False),
        ("units", Unit),
        ("bank_accounts", BankAccount),
        ("tax_schemes", TaxScheme, TaxSchemeMinSerializer, False),
        ("bank_accounts", BankAccount, BankAccountSerializer),
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_sales=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "items",
            Item.objects.filter(
                Q(can_be_purchased=True)
                | Q(direct_expense=True)
                | Q(indirect_expense=True)
                | Q(fixed_asset=True)
            ).select_related("unit"),
            ItemPurchaseSerializer,
        ),
    )

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows", "rows__item", "rows__unit")
        elif self.action == "list":
            qs = qs.select_related("party")
        return qs.order_by("-id")

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.is_issued():
            if not request.company.enable_credit_note_update:
                raise APIException({"detail": "Issued debit notes can't be updated"})
            # self.request.user.check_perm("DebitNoteIssuedModify")
        return super().update(request, *args, **kwargs)

    def get_serializer_class(self):
        if self.action == "list":
            return DebitNoteListSerializer
        elif self.action == "retrieve":
            return DebitNoteDetailSerializer
        return DebitNoteCreateSerializer

    def get_defaults(self, request=None, *args, **kwargs):
        data = {
            "options": {
                "fiscal_years": FiscalYearSerializer(
                    request.company.get_fiscal_years(), many=True
                ).data
            },
            "fields": {"can_update_issued": request.company.enable_debit_note_update},
        }
        return data

    def get_create_defaults(self, request=None, *args, **kwargs):
        options = PurchaseCreateSettingSerializer(request.company.purchase_setting).data
        options["voucher_no"] = get_next_voucher_no(DebitNote, request.company.id)
        return {"options": options}

    def get_update_defaults(self, request=None, *args, **kwargs):
        options = PurchaseUpdateSettingSerializer(request.company.purchase_setting).data
        obj = self.get_object()
        invoice_objs = []
        for inv in obj.invoices.all():
            invoice_objs.append({"id": inv.id, "voucher_no": inv.voucher_no})
        options["purchase_invoice_objs"] = invoice_objs

        if not obj.voucher_no:
            options["voucher_no"] = get_next_voucher_no(
                PurchaseVoucher, request.company.id
            )
        return {"options": options}

    @action(detail=True)
    def details(self, request, pk, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    DebitNoteRow.objects.all()
                    .select_related(
                        "item", "item__category", "unit", "discount_obj", "tax_scheme"
                    )
                    .order_by("pk"),
                )
            )
            .select_related("discount_obj", "bank_account", "party")
        )
        data = DebitNoteDetailSerializer(get_object_or_404(pk=pk, queryset=qs)).data
        data["can_update_issued"] = request.company.enable_debit_note_update
        return Response(data)

    @action(detail=True, methods=["POST"])
    def mark_as_resolved(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        try:
            obj.mark_as_resolved()
            return Response({})
        except Exception as e:
            raise APIException(str(e))

    @action(detail=True, methods=["POST"], url_path="log-print")
    def log_print(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        obj.print_count += 1
        obj.save()
        return Response({"print_count": obj.print_count})

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        debit_note = get_object_or_404(DebitNote, pk=pk)
        journals = debit_note.journal_entries()
        return Response(JournalEntriesSerializer(journals, many=True).data)

    @action(detail=False)
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        params = [
            ("Invoices", queryset, DebitNoteResource),
            (
                "Debit Note Rows",
                DebitNoteRow.objects.filter(voucher__company_id=request.company.id),
                DebitNoteRowResource,
            ),
        ]
        return qs_to_xls(params)


class JournalVoucherViewSet(DeleteRows, CRULViewSet):
    queryset = JournalVoucher.objects.all()
    serializer_class = JournalVoucherCreateSerializer
    model = JournalVoucher
    row = JournalVoucherRow
    # filter_backends = [filters.DjangoFilterBackend, rf_filters.OrderingFilter, rf_filters.SearchFilter,]
    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]

    search_fields = [
        "voucher_no",
        "narration",
    ]
    filterset_class = JournalVoucherFilterSet

    collections = (
        (
            "accounts",
            Account.objects.select_related("category", "parent"),
            AccountSerializer,
            True,
            ["code", "name"],
        ),
    )

    def get_queryset(self, **kwargs):
        qs = super().get_queryset()
        if self.action != "list":
            qs = qs.prefetch_related(
                Prefetch(
                    "rows",
                    JournalVoucherRow.objects.order_by("-type", "id").select_related(
                        "account"
                    ),
                )
            )
        return qs.order_by("-pk")

    def get_serializer_class(self):
        if self.action == "list":
            return JournalVoucherListSerializer
        elif self.action == "retrieve":
            return JournalVoucherDetailSerializer
        return JournalVoucherCreateSerializer

    def get_create_defaults(self, request=None, *args, **kwargs):
        voucher_no = get_next_voucher_no(JournalVoucher, request.company.id)
        data = {
            "fields": {
                "voucher_no": voucher_no,
            }
        }
        return data

    @action(detail=False)
    def get_next_no(self, request, *args, **kwargs):
        voucher_no = get_next_voucher_no(JournalVoucher, request.company.id)
        return Response({"voucher_no": voucher_no})

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk, *args, **kwargs):
        # FIFO inconsistency check
        if (
            request.company.inventory_setting.enable_fifo
            and not request.query_params.get("fifo_inconsistency")
        ):
            raise UnprocessableException(
                detail="This may cause inconsistencies in fifo!",
                code="fifo_inconsistency",
            )

        obj = self.get_object()
        try:
            obj.cancel(reason=request.data.get("message"))
            return Response({})
        except Exception as e:
            raise APIException(str(e))


class InvoiceDesignViewSet(CRULViewSet):
    queryset = InvoiceDesign.objects.all()
    serializer_class = InvoiceDesignSerializer

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        request.data._mutable = True
        design = request.data.pop("design", None)
        serializer = InvoiceDesignSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(**serializer.validated_data)
        if "design" in request.FILES:
            instance.refresh_from_db()
            instance.design = request.FILES.get("design")
            instance.save()
        return Response(serializer.validated_data)

    @action(detail=True)
    def company_invoice(self, request, pk, *args, **kwargs):
        data = {}
        try:
            invoice = InvoiceDesign.objects.get(company_id=pk)
            data = self.serializer_class(invoice).data
        except InvoiceDesign.DoesNotExist:
            pass
        return Response(data)


class SalesDiscountViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = SalesDiscountSerializer

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "name",
    ]
    filterset_class = SalesDiscountFilterSet


class PurchaseDiscountViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = PurchaseDiscountSerializer

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "name",
    ]
    filterset_class = PurchaseDiscountFilterSet


class SalesBookViewSet(
    CompanyViewSetMixin, mixins.ListModelMixin, viewsets.GenericViewSet
):
    serializer_class = SalesBookSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = SalesVoucherFilterSet

    def get_queryset(self, **kwargs):
        qs = (
            super()
            .get_queryset()
            .filter(
                company_id=self.request.company.id,
                status__in=["Issued", "Paid", "Partially Paid"],
            )
            .prefetch_related(
                Prefetch(
                    "rows",
                    SalesVoucherRow.objects.select_related(
                        "discount_obj", "tax_scheme"
                    ),
                )
            )
            .select_related("discount_obj", "party")
        )
        return qs.order_by("-date", "-voucher_no")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        data = serializer.data

        is_filtered = any(
            x in self.request.query_params if self.request.query_params.get(x) else None
            for x in self.filterset_class.base_filters.keys()
        )

        if is_filtered:
            aggregate = queryset.aggregate(
                total__amount=Sum(
                    "total_amount"
                ),  # couldn't use total_amount and total__amount is humanized as total_amount in fe
                total_meta_discount=Sum("meta_discount"),
                total_meta_tax=Sum("meta_tax"),
                total_meta_taxable=Sum("meta_taxable"),
                total_meta_non_taxable=Sum("meta_non_taxable"),
                total_export=Sum(Case(When(is_export=True, then=F("total_amount")))),
            )
            self.paginator.aggregate = aggregate

        return self.get_paginated_response(data)

    @action(detail=False)
    def export(self, request, *args, **kwargs):
        if self.is_filtered_by_date():
            input_stream_path = os.path.join(
                settings.BASE_DIR,
                "store/spreadsheet_templates/np/sales-book.xlsx",
            )
            input_stream = open(input_stream_path, "rb")
            wb = openpyxl.load_workbook(input_stream)
            ws = wb.active
            # TODO Optimization: Check if queryset can be optimized
            qs = (
                super()
                .get_queryset()
                .filter(
                    company_id=self.request.company.id,
                    status__in=["Issued", "Paid", "Partially Paid", "Cancelled"],
                )
                .select_related("discount_obj", "party")
                .order_by("date", "voucher_no")
            )
            queryset = self.filter_queryset(qs).annotate(
                total_quantity=Sum("rows__quantity")
            )
            serializer = SalesBookExportSerializer(queryset, many=True)
            data = serializer.data

            for idx, row in enumerate(data):
                ws.cell(
                    column=1,
                    row=idx + 7,
                    value=ad2bs_str(row.get("date")).replace("-", "."),
                )
                ws.cell(column=2, row=idx + 7, value=row.get("voucher_no"))
                if row.get("status") == "Cancelled":
                    ws.cell(column=3, row=idx + 7, value="Cancelled")
                    ws.cell(column=5, row=idx + 7, value="0")
                    ws.cell(column=6, row=idx + 7, value="0")
                    ws.cell(column=7, row=idx + 7, value="0")
                    ws.cell(column=8, row=idx + 7, value="0")
                else:
                    taxable = row.get("voucher_meta").get("taxable")
                    non_taxable = row.get("voucher_meta").get("non_taxable")
                    ws.cell(column=5, row=idx + 7, value=taxable + non_taxable)
                    ws.cell(column=6, row=idx + 7, value=non_taxable)
                    ws.cell(column=7, row=idx + 7, value=taxable)
                    ws.cell(
                        column=8, row=idx + 7, value=row.get("voucher_meta").get("tax")
                    )
                    ws.cell(column=3, row=idx + 7, value=row.get("buyers_name"))
                    ws.cell(column=4, row=idx + 7, value=row.get("buyers_pan"))

                if row.get("is_export"):
                    ws.cell(column=9, row=idx + 7, value=row.get("grand_total"))

            years = [
                ad2bs(self.request.query_params.get("start_date"))[0],
                ad2bs(self.request.query_params.get("end_date"))[0],
            ]
            years.sort()
            year_str = "/".join(set([str(year) for year in years]))

            ws.cell(
                column=1,
                row=4,
                value="करदाता दर्ता नं (PAN) : {}        करदाताको नाम: {}         साल    {}      कर अवधि: {} -{}".format(
                    self.request.company.tax_identification_number,
                    self.request.company.name,
                    year_str,
                    ad2bs_str(self.request.query_params.get("start_date")).replace(
                        "-", "."
                    ),
                    ad2bs_str(self.request.query_params.get("end_date")).replace(
                        "-", "."
                    ),
                ),
            )

            response = HttpResponse(
                save_virtual_workbook(wb), content_type="application/vnd.ms-excel"
            )
            filename = "{}_{}.xlsx".format("sales-book" + "_", datetime.today().date())
            response["Content-Disposition"] = 'attachment; filename="{}"'.format(
                filename
            )
            return response


class SalesRowViewSet(CompanyViewSetMixin, viewsets.GenericViewSet):
    serializer_class = SalesRowSerializer
    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    filterset_class = SalesRowFilterSet

    search_fields = [
        "voucher__sales_agent__name",
        "voucher__party__name",
        "voucher__party__name",
        "voucher__party__tax_identification_number",
        "item__name",
    ]

    def get_queryset(self, **kwargs):
        qs = SalesVoucherRow.objects.filter(
            voucher__company_id=self.request.company.id
        ).select_related("item", "voucher__party")
        return qs.order_by("-pk")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        data = serializer.data

        is_filtered = any(
            x in self.request.query_params if self.request.query_params.get(x) else None
            for x in self.filterset_class.base_filters.keys()
        )
        if is_filtered:
            aggregate = queryset.aggregate(
                Sum("quantity"),
                Sum("discount_amount"),
                Sum("tax_amount"),
                Sum("net_amount"),
                Avg("rate"),
                Count("item", distinct=True),
                Count("voucher", distinct=True),
                Count("voucher__party", distinct=True),
                Count("voucher__sales_agent", distinct=True),
            )
            self.paginator.aggregate = aggregate
        return self.get_paginated_response(data)

    @action(detail=False, url_path="by-category")
    def by_category(self, request, *args, **kwargs):
        start_date = request.query_params.get("start_date", None)
        end_date = request.query_params.get("end_date", None)
        qs = self.get_queryset()
        if start_date:
            qs = qs.filter(voucher__date__gte=start_date)
        if end_date:
            qs = qs.filter(voucher__date__lte=end_date)
        qs = (
            qs.values("item__category", "item__category__name")
            .annotate(
                quantity=Sum("quantity"),
                discount_amount=Sum("discount_amount"),
                tax_amount=Sum("tax_amount"),
                net_amount=Sum("net_amount"),
            )
            .order_by("-net_amount")
        )
        return Response(qs)


class PurchaseVoucherRowViewSet(CompanyViewSetMixin, viewsets.GenericViewSet):
    serializer_class = PurchaseVoucherRowSerializer
    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    filterset_class = PurchaseVoucherRowFilterSet

    search_fields = [
        "voucher__party__name",
        "voucher__party__tax_identification_number",
        "item__name",
    ]

    def get_queryset(self, **kwargs):
        qs = PurchaseVoucherRow.objects.filter(
            voucher__company_id=self.request.company.id
        ).select_related("item", "voucher__party")
        return qs.order_by("-pk")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        data = serializer.data

        is_filtered = any(
            x in self.request.query_params if self.request.query_params.get(x) else None
            for x in self.filterset_class.base_filters.keys()
        )
        if is_filtered:
            aggregate = queryset.aggregate(
                Sum("quantity"),
                Sum("discount_amount"),
                Sum("tax_amount"),
                Sum("net_amount"),
                Avg("rate"),
                Count("item", distinct=True),
                Count("voucher", distinct=True),
                Count("voucher__party", distinct=True),
                # Count("voucher__sales_agent", distinct=True),
            )
            self.paginator.aggregate = aggregate
        return self.get_paginated_response(data)

    @action(detail=False, url_path="by-category")
    def by_category(self, request, *args, **kwargs):
        start_date = request.query_params.get("start_date", None)
        end_date = request.query_params.get("end_date", None)
        qs = self.get_queryset()
        if start_date:
            qs = qs.filter(voucher__date__gte=start_date)
        if end_date:
            qs = qs.filter(voucher__date__lte=end_date)
        qs = (
            qs.values("item__category", "item__category__name")
            .annotate(
                quantity=Sum("quantity"),
                discount_amount=Sum("discount_amount"),
                tax_amount=Sum("tax_amount"),
                net_amount=Sum("net_amount"),
            )
            .order_by("-net_amount")
        )
        return Response(qs)


class PurchaseBookViewSet(
    CompanyViewSetMixin, mixins.ListModelMixin, viewsets.GenericViewSet
):
    serializer_class = PurchaseBookSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = PurchaseVoucherFilterSet

    def get_queryset(self, **kwargs):
        qs = (
            super()
            .get_queryset()
            .filter(Q(rows__item__can_be_sold=True) | Q(meta_tax__gt=0))
            .filter(
                company_id=self.request.company.id,
                status__in=["Issued", "Paid", "Partially Paid"],
            )
            .prefetch_related(
                Prefetch(
                    "rows",
                    PurchaseVoucherRow.objects.all().select_related(
                        "discount_obj", "tax_scheme"
                    ),
                )
            )
            .select_related("discount_obj", "party")
        )
        return qs.distinct().order_by("-date", "-pk")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        data = serializer.data

        is_filtered = any(
            x in self.request.query_params if self.request.query_params.get(x) else None
            for x in self.filterset_class.base_filters.keys()
        )

        if is_filtered:
            aggregate = queryset.aggregate(
                total_amount=Sum("total_amount"),
                total_meta_discount=Sum("meta_discount"),
                total_meta_tax=Sum("meta_tax"),
                total_meta_taxable=Sum("meta_taxable"),
                total_meta_non_taxable=Sum("meta_non_taxable"),
                # total_import=Sum(Case(When(is_import=True, then=F('total_amount'))))
            )
            self.paginator.aggregate = aggregate

        return self.get_paginated_response(data)

    @action(detail=False)
    def export(self, request, *args, **kwargs):
        if self.is_filtered_by_date():
            input_stream_path = os.path.join(
                settings.BASE_DIR,
                "store/spreadsheet_templates/np/purchase-book.xlsx",
            )
            input_stream = open(input_stream_path, "rb")
            wb = openpyxl.load_workbook(input_stream)
            ws = wb.active
            qs = (
                super()
                .get_queryset()
                .filter(
                    company_id=self.request.company.id,
                    status__in=["Issued", "Paid", "Partially Paid"],
                )
                .select_related("discount_obj", "party")
                .order_by("date", "voucher_no")
            )
            queryset = self.filter_queryset(qs).annotate(
                total_quantity=Sum("rows__quantity")
            )
            serializer = PurchaseBookExportSerializer(queryset, many=True)
            data = serializer.data

            for idx, row in enumerate(data):
                taxable = row.get("voucher_meta").get("taxable")
                non_taxable = row.get("voucher_meta").get("non_taxable")
                ws.cell(
                    column=1,
                    row=idx + 7,
                    value=ad2bs_str(row.get("date")).replace("-", "."),
                )
                ws.cell(column=2, row=idx + 7, value=row.get("voucher_no"))
                ws.cell(column=4, row=idx + 7, value=row.get("party_name"))
                ws.cell(
                    column=5, row=idx + 7, value=row.get("tax_identification_number")
                )
                ws.cell(column=6, row=idx + 7, value=taxable + non_taxable)
                ws.cell(column=7, row=idx + 7, value=non_taxable)
                ws.cell(column=8, row=idx + 7, value=taxable)
                ws.cell(column=9, row=idx + 7, value=row.get("voucher_meta").get("tax"))

            years = [
                ad2bs(self.request.query_params.get("start_date"))[0],
                ad2bs(self.request.query_params.get("end_date"))[0],
            ]
            years.sort()
            year_str = "/".join(set([str(year) for year in years]))

            ws.cell(
                column=1,
                row=4,
                value="करदाता दर्ता नं (PAN) : {}        करदाताको नाम: {}         साल    {}      कर अवधि: {} -{}".format(
                    self.request.company.tax_identification_number,
                    self.request.company.name,
                    year_str,
                    ad2bs_str(self.request.query_params.get("start_date")).replace(
                        "-", "."
                    ),
                    ad2bs_str(self.request.query_params.get("end_date")).replace(
                        "-", "."
                    ),
                ),
            )

            response = HttpResponse(
                save_virtual_workbook(wb), content_type="application/vnd.ms-excel"
            )
            filename = "{}_{}.xlsx".format(
                "purchase-book" + "_", datetime.today().date()
            )
            response["Content-Disposition"] = 'attachment; filename="{}"'.format(
                filename
            )
            return response


class SalesAgentViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = SalesAgentSerializer


class PurchaseSettingsViewSet(CRULViewSet):
    serializer_class = PurchaseSettingCreateSerializer
    collections = (
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_purchase=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "landed_cost_accounts",
            Account.objects.filter(
                category__system_code=settings.ACCOUNT_CATEGORY_SYSTEM_CODES[
                    "Landed Cost"
                ]
            ),
            GenericSerializer,
            True,
            ["name"],
        ),
    )


    def get_defaults(self, request=None, *args, **kwargs):
        p_setting = self.request.company.purchase_setting

        data = {"fields": PurchaseSettingSerializer(p_setting).data}
        return data

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        update_data = request.data
        if not update_data["enable_item_rate_change_alert"]:
            update_data.pop("rate_change_alert_emails")
            instance.update(update_data)
        return super().update(request, *args, **kwargs)


class SalesSettingsViewSet(CRULViewSet):
    serializer_class = SalesSettingCreateSerializer
    collections = (
        (
            "payment_modes",
            PaymentMode.objects.filter(enabled_for_sales=True),
            GenericSerializer,
            True,
            ["name"],
        ),
    )

    def get_defaults(self, request=None, *args, **kwargs):
        s_setting = self.request.company.sales_setting

        data = {
            "fields": SalesSettingsSerializer(
                s_setting, context={"request": request}
            ).data,
            "file_upload_paths": {
                "default_email_attachments": "default_email_attachments",
            },
        }
        return data


class PaymentReceiptViewSet(CRULViewSet):
    serializer_class = PaymentReceiptFormSerializer
    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    queryset = PaymentReceipt.objects.all()
    search_fields = ["party__name", "amount", "cheque_deposit__cheque_number"]

    filterset_class = PaymentReceiptFilterSet
    collections = (("bank_accounts", BankAccount),)

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list":
            qs = qs.select_related("party")
        return qs.order_by("-date", "-id")

    def get_serializer_class(self):
        if self.action == "list":
            return PaymentReceiptSerializer
        return PaymentReceiptFormSerializer

    def get_defaults(self, request=None, *args, **kwargs):
        data = {
            "options": {
                "fiscal_years": FiscalYearSerializer(
                    request.company.get_fiscal_years(), many=True
                ).data
            }
        }
        return data

    @action(detail=True, methods=["POST"])
    def mark_as_cleared(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        obj.clear()
        return Response({})

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk, *args, **kwargs):
        obj = self.get_object()
        obj.cancel()
        return Response({})

    @action(detail=True)
    def details(self, request, pk, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .select_related("bank_account", "cheque_deposit", "party")
        )
        data = PaymentReceiptDetailSerializer(
            get_object_or_404(pk=pk, queryset=qs)
        ).data
        return Response(data)

    @action(detail=False, url_path="fetch-invoice")
    def fetch_invoice(self, request, *args, **kwargs):
        qs = SalesVoucher.objects.filter(company_id=request.company.id).select_related(
            "party"
        )
        invoice = get_object_or_404(
            voucher_no=request.query_params.get("invoice_no"),
            fiscal_year_id=request.query_params.get("fiscal_year"),
            queryset=qs,
        )
        if invoice.status == "Paid":
            return Response(
                {"detail": "Invoice has already been paid for!"}, status=400
            )
        if invoice.status == "Cancelled":
            return Response(
                {"detail": "Invoice has already been canceled!"}, status=400
            )
        if not invoice.party_id:
            return Response(
                {"detail": "Requested invoice isn't for a party!"}, status=400
            )
        data = {
            "id": invoice.id,
            "voucher_no": invoice.voucher_no,
            "party_id": invoice.party_id,
            "party_name": invoice.party.name,
            "amount": invoice.total_amount,
            "taxable": invoice.get_voucher_meta().get("taxable") or 0,
        }
        return Response(data)

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        obj = get_object_or_404(self.get_queryset(), pk=pk)
        journals = obj.journal_entries()
        return Response(JournalEntriesSerializer(journals, many=True).data)

    @action(detail=False, url_path="collection-report")
    def collection_report(self, request, *args, **kwargs):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if not (start_date and end_date):
            raise ValidationError("Start and end dates are required.")

        qs = PaymentReceipt.objects.filter(
            company=request.company, status__in=["Cleared"]
        ).filter(date__gte=start_date, date__lte=end_date)

        excluded_ids = (
            qs.annotate(cnt=Count("invoices__sales_agent_id", distinct=True))
            .filter(cnt__gte=2)
            .values_list("id", flat=True)
        )
        excluded_sum = (
            qs.filter(id__in=excluded_ids).aggregate(Sum("amount")).get("amount__sum")
        )

        included = qs.exclude(id__in=excluded_ids)

        # values = included.values('invoices__sales_agent_id', 'invoices__sales_agent__name').annotate(
        #     total_amount=DistinctSum('amount'))

        values = (
            included.distinct()
            .values("invoices__sales_agent_id", "invoices__sales_agent__name")
            .annotate(total_amount=Sum("amount"))
        )

        return Response(
            {
                "values": values,
                "total": values.aggregate(Sum("amount")).get("amount__sum"),
                "excluded": excluded_ids,
                "excluded_sum": excluded_sum,
            }
        )

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(
            self.get_queryset().prefetch_related("invoices")
        )
        page = self.paginate_queryset(queryset.distinct())
        serializer = self.get_serializer(page, many=True)
        data = serializer.data

        is_filtered = any(
            x in self.request.query_params if self.request.query_params.get(x) else None
            for x in self.filterset_class.base_filters.keys()
        )
        if is_filtered:
            aggregate = queryset.aggregate(Sum("amount"), Sum("tds_amount"))
            aggregate["invoices"] = queryset.aggregate(
                Count("invoices", distinct=True)
            ).get("invoices__count")
            self.paginator.aggregate = aggregate

        return self.get_paginated_response(data)


class ChallanViewSet(InputChoiceMixin, DeleteRows, CRULViewSet):
    queryset = Challan.objects.all()
    serializer_class = ChallanCreateSerializer
    model = Challan
    row = ChallanRow
    collections = [
        ("parties", Party, PartyMinSerializer),
        ("units", Unit),
        (
            "items",
            Item.objects.filter(can_be_sold=True, track_inventory=True).select_related(
                "unit"
            ),
            ItemSalesSerializer,
        ),
    ]

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]

    filterset_class = ChallanFilterSet

    search_fields = [
        "voucher_no",
        "party__name",
        "remarks",
        "party__tax_identification_number",
        "customer_name",
        "rows__item__name",
    ]

    def get_collections(self, request=None, *args, **kwargs):
        sales_agent_tuple = ("sales_agents", SalesAgent)
        if (
            request.company.enable_sales_agents
            and sales_agent_tuple not in self.collections
        ):
            # noinspection PyTypeChecker
            self.collections.append(sales_agent_tuple)
        return super().get_collections(request)

    def get_queryset(self, **kwargs):
        qs = super(ChallanViewSet, self).get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows", "rows__item", "rows__unit")
        elif self.action == "list":
            qs = qs.select_related("party")
        return qs.order_by("-pk")

    def get_serializer_class(self):
        if self.action in ("choices", "list"):
            return ChallanListSerializer
        return ChallanCreateSerializer

    # def update(self, request, *args, **kwargs):
    #     obj = self.get_object()
    #     # if obj.is_issued():
    #     #     if not request.company.enable_sales_invoice_update:
    #     #         raise APIException({'detail': 'Issued sales invoices can\'t be updated'})
    #     #     self.request.user.check_perm('SalesIssuedModify')
    #     return super().update(request, *args, **kwargs)

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        sale_voucher = get_object_or_404(SalesVoucher, pk=pk)
        journals = sale_voucher.journal_entries()
        return Response(SalesJournalEntrySerializer(journals, many=True).data)

    @action(detail=True)
    def details(self, request, pk, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    SalesVoucherRow.objects.all()
                    .select_related("item", "unit", "discount_obj", "tax_scheme")
                    .order_by("pk"),
                )
            )
            .select_related(
                "discount_obj", "bank_account", "company__sales_setting", "party"
            )
        )
        data = SalesVoucherDetailSerializer(
            get_object_or_404(pk=pk, queryset=qs), context={"request": request}
        ).data
        data["can_update_issued"] = request.company.enable_sales_invoice_update
        return Response(data)

    def get_defaults(self, request=None, *args, **kwargs):
        return {
            "options": {"enable_sales_agents": request.company.enable_sales_agents},
        }

    def get_create_defaults(self, request=None, *args, **kwargs):
        data = {
            "options": {"voucher_no": get_next_voucher_no(Challan, request.company.id)}
        }
        return data

    @action(detail=True, methods=["POST"])
    def resolve(self, request, pk, *args, **kwargs):
        challan = self.get_object()
        if challan.status == "Issued":
            challan.status = "Resolved"
            challan.save()
            return Response({})
        return Response({"message": "Cannot resolve the challan."})

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk, *args, **kwargs):
        challan = self.get_object()
        message = request.data.get("message")
        if not message:
            raise RESTValidationError(
                {"message": "message field is required for cancelling invoice!"}
            )

        # FIFO inconsistency check
        if (
            request.company.inventory_setting.enable_fifo
            and not request.query_params.get("fifo_inconsistency")
        ):
            raise UnprocessableException(
                detail="This may cause inconsistencies in fifo!",
                code="fifo_inconsistency",
            )

        challan.cancel(message)
        return Response({})

    @action(detail=True, methods=["POST"], url_path="log-print")
    def log_print(self, request, pk, *args, **kwargs):
        challan = self.get_object()
        challan.print_count += 1
        challan.save()
        return Response({"print_count": challan.print_count})

    @action(detail=False, url_path="by-voucher-no")
    def by_voucher_no(self, request, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows", ChallanRow.objects.all().select_related("item", "unit")
                )
            )
        )
        challan = get_object_or_404(
            voucher_no=request.query_params.get("invoice_no"),
            fiscal_year_id=request.query_params.get("fiscal_year"),
            queryset=qs,
        )
        if challan.sales.exclude(status__iexact="cancelled").exists():
            return Response({"detail": "Challan has already been used."}, status=400)
        if challan.status != "Issued":
            return Response({"detail": "The challan can not be used."}, status=400)
        return Response(ChallanCreateSerializer(challan).data)

        # @action(detail=False, url_path='by-voucher-no')
        # def by_voucher_no(self, request, *args, **kwargs):
        #     qs = super().get_queryset().prefetch_related(
        #         Prefetch('rows',
        #                  SalesVoucherRow.objects.all().select_related('item', 'unit', 'discount_obj',
        #                                                               'tax_scheme'))).select_related(
        #         'discount_obj', 'bank_account')
        #     return Response(
        #         SalesVoucherDetailSerializer(get_object_or_404(voucher_no=request.query_params.get('invoice_no'),
        #                                                        fiscal_year_id=request.query_params.get('fiscal_year'),
        #                                                        queryset=qs)).data)

        # @action(detail=False)
        # def export(self, request, *args, **kwargs):
        # queryset = self.filter_queryset(self.get_queryset())
        #     params = [
        #         ('Invoices', queryset, SalesVoucherResource),
        #         ('Sales Rows', Challan.objects.filter(voucher__company_id=request.company.id),
        #          SalesVoucherRowResource),
        #     ]
        #     return qs_to_xls(params)


class PurchaseOrderViewSet(InputChoiceMixin, DeleteRows, CRULViewSet):
    queryset = PurchaseOrder.objects.all()
    serializer_class = PurchaseOrderCreateSerializer
    model = PurchaseOrder
    row = PurchaseOrderRow

    collections = [
        ("parties", Party, PartyMinSerializer),
        ("units", Unit),
        (
            "items",
            Item.objects.filter(
                can_be_purchased=True, track_inventory=True
            ).select_related("unit"),
            ItemPurchaseSerializer,
        ),
    ]

    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]

    filterset_class = PurchaseOrderFilterSet

    search_fields = [
        "voucher_no",
        "party__name",
        "remarks",
        "party__tax_identification_number",
        "rows__item__name",
    ]

    def get_queryset(self, **kwargs):
        qs = super(PurchaseOrderViewSet, self).get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows")
        elif self.action == "list":
            qs = qs.select_related("party")
        return qs.order_by("-pk")

    def get_serializer_class(self):
        if self.action == "list":
            return PurchaseOrderListSerializer
        return PurchaseOrderCreateSerializer

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk=None, *args, **kwargs):
        instance = self.get_object()
        message = request.data.get("message")
        if not message:
            raise RESTValidationError(
                {"message": "message field is required for cancelling invoice!"}
            )
        instance.remarks = f"Reason for cancellation: {message}"
        instance.status = "Cancelled"
        instance.save()
        return Response({})

    @action(detail=False, url_path="by-voucher-no")
    def by_voucher_no(self, request, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "rows",
                    PurchaseOrderRow.objects.all().select_related("item", "unit"),
                )
            )
        )
        voucher_number = request.query_params.get("invoice_no")
        fiscal_year = request.query_params.get("fiscal_year")
        instance = get_object_or_404(
            voucher_no=voucher_number, fiscal_year_id=fiscal_year, queryset=qs
        )
        if instance.purchases.exclude(status__iexact="cancelled").exists():
            return Response(
                {"detail": "The purchase order has already been used."}, status=400
            )

        if instance.status == "Cancelled":
            return Response(
                {"detail": "The selected purchase order can not be used."}, status=400
            )
        return Response(PurchaseOrderCreateSerializer(instance).data)
