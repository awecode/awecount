from datetime import datetime

from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Q, Sum
from django_filters import rest_framework as filters
from openpyxl import load_workbook
from rest_framework import filters as rf_filters
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.generics import get_object_or_404
from rest_framework.mixins import DestroyModelMixin
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.response import Response

from apps.ledger.models import Account
from apps.ledger.models import Category as AccountCategory
from apps.ledger.models import Transaction as Ledger
from apps.ledger.serializers import AccountMinSerializer
from apps.tax.models import TaxScheme
from apps.tax.serializers import TaxSchemeMinSerializer
from apps.voucher.models import (
    ChallanRow,
    CreditNoteRow,
    DebitNoteRow,
    PurchaseOrderRow,
    PurchaseVoucherRow,
    SalesVoucherRow,
)
from awecount.libs.CustomViewSet import CRULViewSet, GenericSerializer
from awecount.libs.mixins import DeleteRows, InputChoiceMixin, ShortNameChoiceMixin

from .filters import BookFilterSet, InventoryAccountFilterSet, ItemFilterSet
from .models import (
    BillOfMaterial,
    BillOfMaterialRow,
    Brand,
    Category,
    InventoryAccount,
    Item,
    JournalEntry,
    Transaction,
    Unit,
)
from .models import Category as InventoryCategory
from .serializers import (
    BillOfMaterialCreateSerializer,
    BillOfMaterialListSerializer,
    BookSerializer,
    BrandSerializer,
    InventoryAccountSerializer,
    InventoryCategorySerializer,
    InventoryCategoryTrialBalanceSerializer,
    InventorySettingCreateSerializer,
    ItemDetailSerializer,
    ItemListMinSerializer,
    ItemListSerializer,
    ItemOpeningSerializer,
    ItemPOSSerializer,
    ItemSerializer,
    JournalEntrySerializer,
    TransactionEntrySerializer,
    UnitSerializer,
)


class ItemViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = ItemSerializer
    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = [
        "name",
        "code",
        "description",
        "search_data",
        "selling_price",
        "cost_price",
    ]
    filterset_class = ItemFilterSet
    parser_classes = [JSONParser, MultiPartParser]

    collections = (
        ("brands", Brand, BrandSerializer),
        ("inventory_categories", InventoryCategory, InventoryCategorySerializer),
        ("units", Unit, UnitSerializer),
        ("accounts", Account, AccountMinSerializer),
        # ('purchase_accounts', Account.objects.filter(category__name="Purchase"), AccountMinSerializer),
        # ('sales_accounts', Account.objects.filter(category__name="Sales"), AccountMinSerializer),
        ("tax_scheme", TaxScheme, TaxSchemeMinSerializer),
        (
            "discount_allowed_accounts",
            Account.objects.filter(category__name="Discount Expenses"),
            AccountMinSerializer,
        ),
        (
            "discount_received_accounts",
            Account.objects.filter(category__name="Discount Income"),
            AccountMinSerializer,
        ),
    )

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list":
            qs = qs.order_by("-id")
        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return ItemListSerializer
        if self.action == "list_items":
            return ItemListMinSerializer
        return self.serializer_class

    def merge_items(self, item_ids, config=None):
        items = Item.objects.filter(id__in=item_ids)
        for i in range(len(items) - 1):
            item1 = items[i]
            item2 = items[i + 1]
            if (
                (item1.can_be_purchased or item1.can_be_sold or item1.fixed_asset)
                and (item2.direct_expense or item2.indirect_expense)
            ) or (
                (item2.can_be_purchased or item2.can_be_sold or item2.fixed_asset)
                and (item1.direct_expense or item1.indirect_expense)
            ):
                return True

        # Select one item from the items list
        if config and config.get("defaultItem"):
            item = items.get(id=config["defaultItem"])
        else:
            item = items[0]

        remaining_items = items.exclude(id=item.id)

        has_inventory_account = False
        has_purchase_account = False
        has_sales_account = False
        has_discount_received_account = False
        has_discount_allowed_account = False
        # item_account_ids = [x.name for x in items]
        # inventory_account_ids = items.values_list("account", flat=True)
        # inventory_accounts = InventoryAccount.objects.filter(id__in=inventory_account_ids)
        if items.values_list("account").exists():
            has_inventory_account = True
        if items.values_list("purchase_account").exists():
            has_purchase_account = True
        if items.values_list("sales_account").exists():
            has_sales_account = True

        # Set the selected item in purchase rows, sales rows, challan rows, purchase order rows, debit_rows and credit rows
        # if has_inventory_account:
        purchase_voucher_rows = PurchaseVoucherRow.objects.filter(item__id__in=item_ids)
        purchase_voucher_rows.update(item=item)

        sales_voucher_rows = SalesVoucherRow.objects.filter(item__id__in=item_ids)
        sales_voucher_rows.update(item=item)

        challan_rows = ChallanRow.objects.filter(item__id__in=item_ids)
        challan_rows.update(item=item)

        purchase_order_rows = PurchaseOrderRow.objects.filter(item__id__in=item_ids)
        purchase_order_rows.update(item=item)

        credit_note_rows = CreditNoteRow.objects.filter(item__id__in=item_ids)
        credit_note_rows.update(item=item)

        debit_note_rows = DebitNoteRow.objects.filter(item__id__in=item_ids)
        debit_note_rows.update(item=item)

        if has_inventory_account or has_purchase_account or has_sales_account:
            if has_inventory_account:
                if not item.track_inventory:
                    item.track_inventory = True
            if has_purchase_account:
                if not item.can_be_purchased:
                    item.can_be_purchased = True
            if has_sales_account:
                if not item.can_be_sold:
                    item.can_be_sold = True
            item.save()

        remaining_items_inventory_accounts = None

        # Update Inventory account for inventory transactions
        if has_inventory_account:
            inventory_account = item.account

            remaining_items_inventory_account_ids = remaining_items.values_list(
                "account", flat=True
            )
            remaining_items_inventory_accounts = InventoryAccount.objects.filter(
                id__in=remaining_items_inventory_account_ids
            )
            inventory_transactions = Transaction.objects.filter(
                account__in=remaining_items_inventory_accounts
            )

            inventory_transaction_ids = inventory_transactions.values_list(
                "id", flat=True
            )
            journal_entries = JournalEntry.objects.filter(
                transactions__in=inventory_transaction_ids
            )

            for je in journal_entries:
                if je.content_type.name == "item":
                    je.transactions.all().delete()
                    je.delete()

            current_opening_balance = inventory_account.opening_balance

            inventory_transactions.update(account=inventory_account)
            # this assumes that opening_balance_rate of item exixts
            total_quantity = inventory_account.opening_balance + sum(
                obj.opening_balance for obj in remaining_items_inventory_accounts
            )
            if total_quantity:
                inventory_account.opening_balance_rate = (
                    inventory_account.opening_balance / total_quantity
                ) * (inventory_account.opening_balance_rate or 0)
            for obj in remaining_items_inventory_accounts:
                inventory_account.current_balance += obj.current_balance
                inventory_account.opening_balance += obj.opening_balance
                if total_quantity:
                    inventory_account.opening_balance_rate += (
                        obj.opening_balance / total_quantity
                    ) * (obj.opening_balance_rate or 0)
                inventory_account.save()

            if inventory_account.opening_balance != current_opening_balance:
                item.update_opening_balance(item.company.current_fiscal_year)

        if has_purchase_account:
            purchase_account = item.purchase_account
            remaining_items_purchase_account_ids = remaining_items.values_list(
                "purchase_account", flat=True
            )
            remaining_items_purchase_accounts = Account.objects.filter(
                id__in=remaining_items_purchase_account_ids
            )
            purchase_transactions = Ledger.objects.filter(
                account__in=remaining_items_purchase_accounts
            )
            # purchase_transactions_ids = purchase_transactions.values_list("id", flat=True)
            # journals = AccountJournal.objects.filter(transactions__in=purchase_transactions_ids)
            # for je in journals:
            #     if je.content_type.name == "purchase voucher row":
            #         je.transactions.all().delete()
            #         je.delete()
            purchase_transactions.update(account=purchase_account)
            remaining_items_purchase_accounts.delete()

            discount_received_account = item.discount_received_account
            remaining_items_discount_received_account_ids = remaining_items.values_list(
                "discount_received_account", flat=True
            )
            remaining_items_discount_received_accounts = Account.objects.filter(
                id__in=remaining_items_discount_received_account_ids
            )
            discount_received_transactions = Ledger.objects.filter(
                account__in=remaining_items_discount_received_accounts
            )
            discount_received_transactions.update(account=discount_received_account)
            remaining_items_discount_received_accounts.delete()

        if has_sales_account:
            sales_account = item.sales_account
            remaining_items_sales_account_ids = remaining_items.values_list(
                "sales_account", flat=True
            )
            remaining_items_sales_accounts = Account.objects.filter(
                id__in=remaining_items_sales_account_ids
            )
            sales_transactions = Ledger.objects.filter(
                account__in=remaining_items_sales_accounts
            )
            # sales_transactions_ids = sales_transactions.values_list("id", flat=True)
            # journals = AccountJournal.objects.filter(transactions__in=sales_transactions_ids)
            # for je in journals:
            #     if je.content_type.name == "sales voucher row":
            #         je.transactions.all().delete()
            #         je.delete()
            sales_transactions.update(account=sales_account)
            remaining_items_sales_accounts.delete()

            discount_allowed_account = item.discount_allowed_account
            remaining_items_discount_allowed_account_ids = remaining_items.values_list(
                "discount_allowed_account", flat=True
            )
            remaining_items_discount_allowed_accounts = Account.objects.filter(
                id__in=remaining_items_discount_allowed_account_ids
            )
            discount_allowed_transactions = Ledger.objects.filter(
                account__in=remaining_items_discount_allowed_accounts
            )
            discount_allowed_transactions.update(account=discount_allowed_account)
            remaining_items_discount_allowed_accounts.delete()

        # Delete other items
        if remaining_items_inventory_accounts:
            remaining_items_inventory_accounts.delete()
        remaining_items.delete()
        return False

    @action(detail=False, url_path="list")
    def list_items(self, request):
        qs = super().get_queryset().order_by("name")
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, url_path="similar-items")
    def similar_items(self, request):
        from thefuzz import fuzz

        # TODO Use postgres trigram/cosine/levenshtein distance
        qs = super().get_queryset()
        items = qs.values_list("id", "name", "code")
        res = []
        for item in items:
            obj = {}
            similar_items = []
            for id, name, code in items:
                if fuzz.ratio(item[1], name) > 80:
                    similar_items.append(
                        {
                            "id": id,
                            "name": f"{name} ({code})",
                            "code": code,
                            # "match": fuzz.ratio(item[1], name)
                        }
                    )
            if len(similar_items) > 1:
                obj["items"] = similar_items
                obj["config"] = {}
                # res[item[1]] = sim
                res.append(obj)

        unique_ids = {}

        # Filter out duplicate items
        filtered_data = []
        index = 0
        for group in res:
            unique_items = []
            for item in group["items"]:
                item_id = item["id"]
                if item_id not in unique_ids:
                    unique_ids[item_id] = True
                    unique_items.append(item)
            if unique_items:
                index += 1
                filtered_data.append(
                    {
                        "items": [x["id"] for x in unique_items],
                        "config": group["config"],
                        "index": index,
                    }
                )

        return Response(filtered_data)

    @action(detail=False, methods=["POST"])
    def merge(self, request):
        groups_not_merged = []
        items_not_merged = []
        flag = False
        for index, group in enumerate(request.data):
            if group.get("config"):
                ret = self.merge_items(group["items"], group["config"])
            else:
                ret = self.merge_items(group["items"])
            if ret:
                flag = True
                groups_not_merged.append(index + 1)
                items_not_merged.append(group)
        if flag:
            res = {
                "error": {
                    "message": f"Items in Groups {','.join([str(x) for x in groups_not_merged])} were not merged due to conflicting config on items.",
                    "items": items_not_merged,
                }
            }
            return Response(res, status=209)
        return Response(status=200)

    @action(detail=True)
    def details(self, request, pk=None):
        qs = (
            super()
            .get_queryset()
            .select_related(
                "account",
                "sales_account",
                "purchase_account",
                "discount_allowed_account",
                "discount_received_account",
                "expense_account",
                "fixed_asset_account",
                "tax_scheme",
            )
        )
        item = get_object_or_404(queryset=qs, pk=pk)
        serializer = ItemDetailSerializer(item, context={"request": request}).data
        return Response(serializer)

    # items listing for POS
    @action(detail=False)
    def pos(self, request):
        self.filter_backends = (rf_filters.SearchFilter,)
        self.queryset = self.get_queryset().filter(can_be_sold=True)
        self.serializer_class = ItemPOSSerializer
        self.search_fields = [
            "name",
            "code",
            "description",
            "search_data",
        ]
        self.paginator.page_size = settings.POS_ITEMS_SIZE
        return super().list(request)

    @action(detail=False, url_path="sales-choices")
    def sales_choices(self, request):
        queryset = self.get_queryset().filter(can_be_sold=True)
        serializer = GenericSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["POST"], url_path="import")
    def import_from_file(self, request):
        file = request.FILES["file"]
        wb = load_workbook(file)
        ws = wb.active
        ws.delete_rows(1)
        items = []

        def get_bool_for_txt(value):
            if not value:
                return False
            elif value.upper() == "F":
                return False
            elif value.upper() == "T":
                return True
            else:
                raise ValueError("Invalid value detected.")

        def get_category(value):
            if not value:
                return None
            else:
                qs = Category.objects.filter(
                    name__iexact=value, company_id=request.company.id
                )
                if qs.exists():
                    return qs.first()
                else:
                    if request.data.get("create_new_category") == "true":
                        cat = Category.objects.create(
                            name=value, company_id=request.company.id
                        )
                        return cat
                    return None

        def get_Tax(value):
            if not value:
                return None
            else:
                qs = TaxScheme.objects.filter(
                    name__iexact=value, company_id=request.company.id
                )
                if qs.exists():
                    return qs.first()
                else:
                    return None

        def get_unit(value):
            if not value:
                return None
            else:
                qs = Unit.objects.filter(
                    name__iexact=value, company_id=request.company.id
                )
                if qs.exists():
                    return qs.first()
                else:
                    if request.data.get("create_new_unit") == "true":
                        new_unit = Unit.objects.create(
                            name=value, company_id=request.company.id
                        )
                        return new_unit
                    return None

        for row in ws.iter_rows(values_only=True):
            item = {
                "company_id": request.company.id,
                "name": row[0],
                "code": row[1],
                "category": get_category(row[2]),
                "cost_price": row[3],
                "selling_price": row[4],
                "can_be_purchased": get_bool_for_txt(row[5]),
                "can_be_sold": get_bool_for_txt(row[6]),
                "track_inventory": get_bool_for_txt(row[7]),
                "tax_scheme": get_Tax(row[8]),
                "unit": get_unit(row[9]),
            }
            item_obj = Item(**item)
            items.append(item_obj)

        with transaction.atomic():
            try:
                Item.objects.bulk_create(items, batch_size=1000)
                # items = Item.objects.bulk_create(items)
                accounts_to_create = []
                items_to_update = []
                for i, item in enumerate(items):
                    # print(i)
                    if item.can_be_purchased:
                        name = item.name + " (Purchase)"
                        purchase_account = Account(name=name, company=item.company)
                        purchase_account.add_category("Purchase")
                        purchase_account.suggest_code(item)
                        accounts_to_create.append(purchase_account)
                        item.purchase_account = purchase_account

                        name = "Discount Received - " + item.name
                        discount_received_acc = Account(name=name, company=item.company)
                        discount_received_acc.add_category("Discount Income")
                        discount_received_acc.suggest_code(item)
                        accounts_to_create.append(discount_received_acc)
                        item.discount_received_account = discount_received_acc

                    if item.can_be_sold:
                        name = item.name + " (Sales)"
                        sales_account = Account(name=name, company=item.company)
                        sales_account.add_category("Sales")
                        sales_account.suggest_code(item)
                        accounts_to_create.append(sales_account)
                        item.sales_account = sales_account

                        name = "Discount Allowed - " + item.name
                        discount_allowed_account = Account(
                            name=name, company=item.company
                        )
                        discount_allowed_account.add_category("Discount Expenses")
                        discount_allowed_account.suggest_code(item)
                        accounts_to_create.append(discount_allowed_account)
                        item.discount_allowed_account = discount_allowed_account

                    items_to_update.append(item)

                Account.objects.bulk_create(accounts_to_create, batch_size=1000)
                Item.objects.bulk_update(
                    items_to_update,
                    fields=[
                        "purchase_account",
                        "sales_account",
                        "discount_received_account",
                        "discount_allowed_account",
                    ],
                )

            except IntegrityError as e:
                code = e.args[0].split("=(")[1].split(")")[0].split(",")[0]
                res_msg = f"Duplicate items with code {code} detected."
                return Response({"details": res_msg}, status=400)
        return Response({}, status=200)

    @action(detail=True, methods=["GET"], url_path="available-stock")
    def available_stock_data(self, request, pk=None):
        item = self.get_object()
        return Response(item.available_stock_data, status=200)


class ItemOpeningBalanceViewSet(DestroyModelMixin, CRULViewSet):
    serializer_class = ItemOpeningSerializer
    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = [
        "item__name",
        "item__code",
        "item__description",
        "item__search_data",
        "opening_balance",
    ]
    filterset_class = InventoryAccountFilterSet

    collections = (
        (
            "items",
            Item.objects.only("id", "name").filter(
                track_inventory=True, account__opening_balance=0
            ),
            GenericSerializer,
        ),
    )

    def get_queryset(self, company_id=None):
        return super().get_queryset().filter(opening_balance__gt=0)

    def get_update_defaults(self, request=None):
        self.collections = (
            (
                "items",
                Item.objects.filter(track_inventory=True).filter(
                    Q(account__opening_balance=0) | Q(account_id=self.kwargs.get("pk"))
                ),
                ItemListSerializer,
            ),
        )
        return self.get_defaults(request=request)

    def create(self, request, *args, **kwargs):
        data = request.data
        account = get_object_or_404(
            InventoryAccount,
            item__id=data.get("item_id"),
            opening_balance=0,
            company=request.company,
        )
        opening_balance = data.get("opening_balance", None)
        if not opening_balance:
            raise ValidationError({"opening_balance": ["Opening balance is required."]})
        account.opening_balance = data.get("opening_balance")
        account.opening_balance_rate = data.get("opening_balance_rate")
        account.save()
        fiscal_year = self.request.company.current_fiscal_year
        account.item.update_opening_balance(fiscal_year)
        return Response({})

    def perform_update(self, serializer):
        super().perform_update(serializer)
        fiscal_year = self.request.company.current_fiscal_year
        serializer.instance.item.update_opening_balance(fiscal_year)

    def destroy(self, request, *args, **kwargs):
        account = self.get_object()
        account.opening_balance = 0
        account.opening_balance_rate = 0
        account.save()
        fiscal_year = self.request.company.current_fiscal_year
        account.item.update_opening_balance(fiscal_year)
        JournalEntry.objects.filter(
            content_type__model="item",
            content_type__app_label="product",
            object_id=account.item.id,
        ).delete()
        return Response({})


class BookViewSet(InputChoiceMixin, CRULViewSet):
    collections = (("brands", Brand, BrandSerializer),)

    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = [
        "name",
        "code",
        "description",
        "search_data",
        "selling_price",
        "cost_price",
    ]
    filterset_class = BookFilterSet

    def get_queryset(self, **kwargs):
        queryset = Item.objects.filter(
            category__name="Book", company=self.request.company
        )
        return queryset

    serializer_class = BookSerializer

    @action(detail=False)
    def category(self, request):
        cat = Category.objects.filter(company=self.request.company, name="Book").first()
        return Response(InventoryCategorySerializer(cat).data)


class UnitViewSet(InputChoiceMixin, ShortNameChoiceMixin, CRULViewSet):
    serializer_class = UnitSerializer


class InventoryCategoryViewSet(InputChoiceMixin, ShortNameChoiceMixin, CRULViewSet):
    serializer_class = InventoryCategorySerializer
    collections = (
        ("units", Unit, UnitSerializer),
        ("accounts", Account, AccountMinSerializer),
        # ('purchase_accounts', Account.objects.filter(category__name="Purchase"), AccountMinSerializer),
        # ('sales_accounts', Account.objects.filter(category__name="Sales"), AccountMinSerializer),
        ("tax_scheme", TaxScheme, TaxSchemeMinSerializer),
        (
            "discount_allowed_accounts",
            Account.objects.filter(category__name="Discount Expenses"),
            AccountMinSerializer,
        ),
        (
            "discount_received_accounts",
            Account.objects.filter(category__name="Discount Income"),
            AccountMinSerializer,
        ),
    )

    def get_collections(self, request=None):
        collections_data = super().get_collections(self.request)
        collections_data["fixed_assets_categories"] = GenericSerializer(
            AccountCategory.objects.get(
                name="Fixed Assets", default=True, company=self.request.company
            ).get_descendants(include_self=True),
            many=True,
        ).data
        collections_data["direct_expenses_categories"] = GenericSerializer(
            AccountCategory.objects.get(
                name="Direct Expenses", default=True, company=self.request.company
            ).get_descendants(include_self=True),
            many=True,
        ).data
        collections_data["indirect_expenses_categories"] = GenericSerializer(
            AccountCategory.objects.get(
                name="Indirect Expenses", default=True, company=self.request.company
            ).get_descendants(include_self=True),
            many=True,
        ).data
        return collections_data

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list":
            qs = qs.order_by("-id")
        return qs

    @action(detail=False, url_path="trial-balance")
    def trial_balance(self, request):
        qs = self.get_queryset().filter(Q(track_inventory=True) | Q(fixed_asset=True))
        return Response(InventoryCategoryTrialBalanceSerializer(qs, many=True).data)


class BrandViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = BrandSerializer


class InventoryAccountViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = InventoryAccountSerializer

    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = (
        "code",
        "name",
    )

    def get_account_ids(self, obj):
        return [obj.id]

    @action(detail=True, methods=["get"], url_path="journal-entries")
    def journal_entries(self, request, pk=None):
        param = request.GET
        start_date = param.get("start_date")
        end_date = param.get("end_date")
        obj = self.get_object()
        entries = (
            JournalEntry.objects.filter(transactions__account_id=obj.pk)
            .order_by("pk", "date")
            .prefetch_related("transactions", "content_type", "transactions__account")
            .select_related()
        )

        if start_date or end_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.strptime(end_date, "%Y-%m-%d")

            if start_date == end_date:
                entries = entries.filter(date=start_date)
            else:
                entries = entries.filter(date__range=[start_date, end_date])
        serializer = JournalEntrySerializer(
            entries, context={"account": obj}, many=True
        )
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def transactions(self, request, pk=None):
        param = request.GET
        obj = self.get_object()
        serializer_class = self.get_serializer_class()
        data = serializer_class(obj).data
        account_ids = self.get_account_ids(obj)
        start_date = param.get("start_date", None)
        end_date = param.get("end_date", None)
        transactions = (
            Transaction.objects.filter(account_id__in=account_ids)
            .order_by("-journal_entry__date", "-pk")
            .select_related("journal_entry__content_type")
        )

        aggregate = {}
        if start_date or end_date:
            if start_date == "null":
                from rest_framework.exceptions import ValidationError

                raise ValidationError(
                    "Either start date or both dates are required to filter."
                )
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
            # TODO: if only start date is given, raise error or process some other way?
            end_date = (
                datetime.strptime(end_date, "%Y-%m-%d")
                if end_date != "null"
                else datetime.today()
            )
            if start_date == end_date:
                transactions = transactions.filter(journal_entry__date=start_date)
            else:
                transactions = transactions.filter(
                    journal_entry__date__range=[start_date, end_date]
                )
            aggregate = transactions.aggregate(Sum("dr_amount"), Sum("cr_amount"))

        # Only show 5 because fetching voucher_no is expensive because of GFK, GFK to be cached
        # self.paginator.page_size = 5
        page = self.paginate_queryset(transactions)
        serializer = TransactionEntrySerializer(page, many=True)
        data["transactions"] = self.paginator.get_response_data(serializer.data)
        data["aggregate"] = aggregate
        return Response(data)

    @action(detail=False, url_path="trial-balance")
    def trial_balance(self, request, format=None):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        if start_date and end_date:
            qq = (
                InventoryAccount.objects.filter(company=request.company)
                .annotate(
                    od=Sum(
                        "transactions__dr_amount",
                        filter=Q(transactions__journal_entry__date__lt=start_date),
                    ),
                    oc=Sum(
                        "transactions__cr_amount",
                        filter=Q(transactions__journal_entry__date__lt=start_date),
                    ),
                    cd=Sum(
                        "transactions__dr_amount",
                        filter=Q(transactions__journal_entry__date__lte=end_date),
                    ),
                    cc=Sum(
                        "transactions__cr_amount",
                        filter=Q(transactions__journal_entry__date__lte=end_date),
                    ),
                )
                .values("id", "name", "item__category_id", "od", "oc", "cd", "cc")
                .exclude(od=None, oc=None, cd=None, cc=None)
            )
            return Response(list(qq))
        return Response({})


class InventorySettingsViewSet(CRULViewSet):
    serializer_class = InventorySettingCreateSerializer

    def get_defaults(self, request=None):
        i_setting = self.request.company.inventory_setting
        data = {"fields": self.get_serializer(i_setting).data}
        return data


class BillOfMaterialViewSet(DeleteRows, CRULViewSet):
    model = BillOfMaterial
    row = BillOfMaterialRow
    serializer_class = BillOfMaterialCreateSerializer
    collections = [
        [
            "finished_products",
            Item.objects.only("id", "name").filter(
                track_inventory=True, bill_of_material__isnull=True
            ),
            GenericSerializer,
        ],
        ["units", Unit],
        [
            "items",
            Item.objects.only("id", "name").filter(track_inventory=True),
            GenericSerializer,
        ],
    ]
    filter_backends = [
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "finished_product__name",
    ]

    def get_queryset(self, **kwargs):
        qs = super(BillOfMaterialViewSet, self).get_queryset()
        return qs.order_by(
            "-created_at",
        )

    def get_serializer_class(self):
        if self.action == "list":
            return BillOfMaterialListSerializer
        return BillOfMaterialCreateSerializer
