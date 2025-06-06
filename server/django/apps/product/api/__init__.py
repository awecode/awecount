from collections import defaultdict
from datetime import datetime
from typing import Literal

from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django_filters import rest_framework as filters
from openpyxl import load_workbook
from rest_framework import filters as rf_filters
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.generics import get_object_or_404
from rest_framework.mixins import DestroyModelMixin
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.response import Response
from xlsxwriter import Workbook

from apps.ledger.models import Account
from apps.ledger.models import Category as AccountCategory
from apps.ledger.models import Transaction as Ledger
from apps.ledger.serializers import AccountMinSerializer, SalesJournalEntrySerializer
from apps.tax.models import TaxScheme
from apps.tax.serializers import TaxSchemeMinSerializer
from apps.voucher.models import (
    ChallanRow,
    CreditNoteRow,
    DebitNoteRow,
    PurchaseOrderRow,
    PurchaseVoucherRow,
    SalesVoucherRow,
)
from awecount.libs.CustomViewSet import CRULViewSet, GenericSerializer
from awecount.libs.exception import UnprocessableException
from awecount.libs.mixins import DeleteRows, InputChoiceMixin, ShortNameChoiceMixin

from ..filters import (
    BookFilterSet,
    InventoryAccountFilterSet,
    InventoryAdjustmentVoucherFilterSet,
    InventoryConversionVoucherFilterSet,
    ItemFilterSet,
)
from ..models import (
    BillOfMaterial,
    BillOfMaterialRow,
    Brand,
    Category,
    InventoryAccount,
    InventoryAdjustmentVoucher,
    InventoryAdjustmentVoucherRow,
    # InventoryConversionVoucherRow,
    InventoryConversionVoucher,
    Item,
    JournalEntry,
    Transaction,
    Unit,
)
from ..models import Category as InventoryCategory
from ..serializers import (
    BillOfMaterialCreateSerializer,
    BillOfMaterialListSerializer,
    BookSerializer,
    BrandSerializer,
    InventoryAccountSerializer,
    InventoryAdjustmentVoucherCreateSerializer,
    InventoryAdjustmentVoucherDetailSerializer,
    InventoryAdjustmentVoucherListSerializer,
    InventoryCategoryFormSerializer,
    InventoryCategorySerializer,
    InventoryCategoryTrialBalanceSerializer,
    InventoryConversionVoucherCreateSerializer,
    InventoryConversionVoucherDetailSerializer,
    InventoryConversionVoucherListSerializer,
    InventorySettingCreateSerializer,
    ItemDetailSerializer,
    ItemFormSerializer,
    ItemListMinSerializer,
    ItemListSerializer,
    ItemOpeningSerializer,
    ItemPOSSerializer,
    ItemSerializer,
    JournalEntrySerializer,
    TransactionEntrySerializer,
    UnitSerializer,
)
from ..serializers.stock_movement import StockMovementSerializer

acc_cat_system_codes = settings.ACCOUNT_CATEGORY_SYSTEM_CODES


class ItemViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = ItemSerializer
    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = [
        "name",
        "code",
        "description",
        "search_data",
        "selling_price",
        "cost_price",
    ]
    filterset_class = ItemFilterSet
    parser_classes = [JSONParser, MultiPartParser]

    collections = (
        ("brands", Brand, BrandSerializer, True, ["name"]),
        (
            "inventory_categories",
            InventoryCategory.objects.select_related(
                "default_unit",
                "sales_account",
                "purchase_account",
                "discount_allowed_account",
                "discount_received_account",
            ),
            InventoryCategoryFormSerializer,
            True,
            ["name"],
        ),
        ("units", Unit, UnitSerializer, True, ["name", "short_name"]),
        ("accounts", Account, AccountMinSerializer, True, ["name"]),
        # ('purchase_accounts', Account.objects.filter(category__name="Purchase"), AccountMinSerializer),
        # ('sales_accounts', Account.objects.filter(category__name="Sales"), AccountMinSerializer),
        ("tax_scheme", TaxScheme, TaxSchemeMinSerializer, True, ["name"]),
        (
            "discount_allowed_accounts",
            Account.objects.filter(
                category__system_code=acc_cat_system_codes["Discount Expenses"]
            ),
            AccountMinSerializer,
            True,
            ["name"],
        ),
        (
            "discount_received_accounts",
            Account.objects.filter(
                category__system_code=acc_cat_system_codes["Discount Income"]
            ),
            AccountMinSerializer,
            True,
            ["name"],
        ),
    )

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list":
            qs = qs.order_by("-id").select_related("category")
        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return ItemListSerializer
        if self.action == "list_items":
            return ItemListMinSerializer
        if self.action == "retrieve":
            return ItemFormSerializer
        return self.serializer_class

    def get_defaults(self, request=None):
        acc_system_codes = settings.ACCOUNT_SYSTEM_CODES
        system_codes = [
            acc_system_codes["Sales Account"],
            acc_system_codes["Purchase Account"],
            acc_system_codes["Discount Expenses"],
            acc_system_codes["Discount Income"],
        ]
        accounts = Account.objects.filter(
            company=request.company, system_code__in=system_codes
        ).values_list("system_code", "id")
        account_ids = {system_code: id for system_code, id in accounts}
        return {
            "options": {
                "global_accounts": {
                    "sales_account_id": account_ids.get(
                        acc_system_codes["Sales Account"]
                    ),
                    "purchase_account_id": account_ids.get(
                        acc_system_codes["Purchase Account"]
                    ),
                    "discount_allowed_account_id": account_ids.get(
                        acc_system_codes["Discount Expenses"]
                    ),
                    "discount_received_account_id": account_ids.get(
                        acc_system_codes["Discount Income"]
                    ),
                }
            },
        }

    def merge_items(self, item_ids, config=None):
        items = Item.objects.filter(id__in=item_ids)
        for i in range(len(items) - 1):
            item1 = items[i]
            item2 = items[i + 1]
            if (
                (item1.can_be_purchased or item1.can_be_sold or item1.fixed_asset)
                and (item2.direct_expense or item2.indirect_expense)
            ) or (
                (item2.can_be_purchased or item2.can_be_sold or item2.fixed_asset)
                and (item1.direct_expense or item1.indirect_expense)
            ):
                return True

        # Select one item from the items list
        if config and config.get("defaultItem"):
            item = items.get(id=config["defaultItem"])
        else:
            item = items[0]

        remaining_items = items.exclude(id=item.id)

        has_inventory_account = False
        has_purchase_account = False
        has_sales_account = False
        has_discount_received_account = False
        has_discount_allowed_account = False
        # item_account_ids = [x.name for x in items]
        # inventory_account_ids = items.values_list("account", flat=True)
        # inventory_accounts = InventoryAccount.objects.filter(id__in=inventory_account_ids)
        if items.values_list("account").exists():
            has_inventory_account = True
        if items.values_list("purchase_account").exists():
            has_purchase_account = True
        if items.values_list("sales_account").exists():
            has_sales_account = True

        # Set the selected item in purchase rows, sales rows, challan rows, purchase order rows, debit_rows and credit rows
        # if has_inventory_account:
        purchase_voucher_rows = PurchaseVoucherRow.objects.filter(item__id__in=item_ids)
        purchase_voucher_rows.update(item=item)

        sales_voucher_rows = SalesVoucherRow.objects.filter(item__id__in=item_ids)
        sales_voucher_rows.update(item=item)

        challan_rows = ChallanRow.objects.filter(item__id__in=item_ids)
        challan_rows.update(item=item)

        purchase_order_rows = PurchaseOrderRow.objects.filter(item__id__in=item_ids)
        purchase_order_rows.update(item=item)

        credit_note_rows = CreditNoteRow.objects.filter(item__id__in=item_ids)
        credit_note_rows.update(item=item)

        debit_note_rows = DebitNoteRow.objects.filter(item__id__in=item_ids)
        debit_note_rows.update(item=item)

        inventory_adjustment_rows = InventoryAdjustmentVoucherRow.objects.filter(
            item__id__in=item_ids
        )
        inventory_adjustment_rows.update(item=item)

        # inventory_conversion_rows = InventoryConversionVoucherRow.objects.filter(
        #     item__id__in=item_ids
        # )
        # inventory_conversion_rows.update(item=item)

        if has_inventory_account or has_purchase_account or has_sales_account:
            if has_inventory_account:
                if not item.track_inventory:
                    item.track_inventory = True
            if has_purchase_account:
                if not item.can_be_purchased:
                    item.can_be_purchased = True
            if has_sales_account:
                if not item.can_be_sold:
                    item.can_be_sold = True
            item.save()

        remaining_items_inventory_accounts = None

        # Update Inventory account for inventory transactions
        if has_inventory_account:
            inventory_account = item.account

            remaining_items_inventory_account_ids = remaining_items.values_list(
                "account", flat=True
            )
            remaining_items_inventory_accounts = InventoryAccount.objects.filter(
                id__in=remaining_items_inventory_account_ids
            )
            inventory_transactions = Transaction.objects.filter(
                account__in=remaining_items_inventory_accounts
            )

            inventory_transaction_ids = inventory_transactions.values_list(
                "id", flat=True
            )
            journal_entries = JournalEntry.objects.filter(
                transactions__in=inventory_transaction_ids
            )

            for je in journal_entries:
                if je.content_type.name == "item":
                    je.transactions.all().delete()
                    je.delete()

            current_opening_balance = inventory_account.opening_balance

            inventory_transactions.update(account=inventory_account)
            # this assumes that opening_balance_rate of item exixts
            total_quantity = inventory_account.opening_balance + sum(
                obj.opening_balance for obj in remaining_items_inventory_accounts
            )
            if total_quantity:
                inventory_account.opening_balance_rate = (
                    inventory_account.opening_balance / total_quantity
                ) * (inventory_account.opening_balance_rate or 0)
            for obj in remaining_items_inventory_accounts:
                inventory_account.current_balance += obj.current_balance
                inventory_account.opening_balance += obj.opening_balance
                if total_quantity:
                    inventory_account.opening_balance_rate += (
                        obj.opening_balance / total_quantity
                    ) * (obj.opening_balance_rate or 0)
                inventory_account.save()

            if inventory_account.opening_balance != current_opening_balance:
                item.update_opening_balance(item.company.current_fiscal_year)

        if has_purchase_account:
            purchase_account = item.purchase_account
            remaining_items_purchase_account_ids = remaining_items.values_list(
                "purchase_account", flat=True
            )
            remaining_items_dedicated_purchase_account_ids = (
                remaining_items.values_list("dedicated_purchase_account", flat=True)
            )
            remaining_items_dedicated_purchase_accounts = Account.objects.filter(
                id__in=remaining_items_dedicated_purchase_account_ids
            )
            remaining_items_purchase_accounts = Account.objects.filter(
                id__in=remaining_items_purchase_account_ids
            )
            purchase_transactions = Ledger.objects.filter(
                account__in=remaining_items_purchase_accounts
            )
            # purchase_transactions_ids = purchase_transactions.values_list("id", flat=True)
            # journals = AccountJournal.objects.filter(transactions__in=purchase_transactions_ids)
            # for je in journals:
            #     if je.content_type.name == "purchase voucher row":
            #         je.transactions.all().delete()
            #         je.delete()
            purchase_transactions.update(account=purchase_account)
            remaining_items_dedicated_purchase_accounts.delete()

            discount_received_account = item.discount_received_account
            remaining_items_discount_received_account_ids = remaining_items.values_list(
                "discount_received_account", flat=True
            )
            remaining_items_dedicated_discount_received_account_ids = (
                remaining_items.values_list(
                    "dedicated_discount_received_account", flat=True
                )
            )
            remaining_items_dedicated_discount_received_accounts = (
                Account.objects.filter(
                    id__in=remaining_items_dedicated_discount_received_account_ids
                )
            )
            remaining_items_discount_received_accounts = Account.objects.filter(
                id__in=remaining_items_discount_received_account_ids
            )
            discount_received_transactions = Ledger.objects.filter(
                account__in=remaining_items_discount_received_accounts
            )
            discount_received_transactions.update(account=discount_received_account)
            remaining_items_dedicated_discount_received_accounts.delete()

        if has_sales_account:
            sales_account = item.sales_account
            remaining_items_sales_account_ids = remaining_items.values_list(
                "sales_account", flat=True
            )
            remaining_items_sales_accounts = Account.objects.filter(
                id__in=remaining_items_sales_account_ids
            )
            remaining_items_dedicated_sales_accounts_ids = remaining_items.values_list(
                "dedicated_sales_account", flat=True
            )
            remaining_items_dedicated_sales_accounts = Account.objects.filter(
                id__in=remaining_items_dedicated_sales_accounts_ids
            )
            sales_transactions = Ledger.objects.filter(
                account__in=remaining_items_sales_accounts
            )
            # sales_transactions_ids = sales_transactions.values_list("id", flat=True)
            # journals = AccountJournal.objects.filter(transactions__in=sales_transactions_ids)
            # for je in journals:
            #     if je.content_type.name == "sales voucher row":
            #         je.transactions.all().delete()
            #         je.delete()
            sales_transactions.update(account=sales_account)
            remaining_items_dedicated_sales_accounts.delete()

            discount_allowed_account = item.discount_allowed_account
            remaining_items_discount_allowed_account_ids = remaining_items.values_list(
                "discount_allowed_account", flat=True
            )
            remaining_items_discount_allowed_accounts = Account.objects.filter(
                id__in=remaining_items_discount_allowed_account_ids
            )
            remaining_items_dedicated_discount_allowed_accounts_ids = (
                remaining_items.values_list(
                    "dedicated_discount_allowed_account", flat=True
                )
            )
            remaining_items_dedicated_discount_allowed_accounts = (
                Account.objects.filter(
                    id__in=remaining_items_dedicated_discount_allowed_accounts_ids
                )
            )
            discount_allowed_transactions = Ledger.objects.filter(
                account__in=remaining_items_discount_allowed_accounts
            )
            discount_allowed_transactions.update(account=discount_allowed_account)
            remaining_items_dedicated_discount_allowed_accounts.delete()

        # Delete other items
        if remaining_items_inventory_accounts:
            remaining_items_inventory_accounts.delete()
        remaining_items.delete()
        return False

    @action(detail=False, url_path="list")
    def list_items(self, request, *args, **kwargs):
        qs = super().get_queryset().order_by("name")
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, url_path="similar-items")
    def similar_items(self, request, *args, **kwargs):
        from thefuzz import fuzz

        # TODO Use postgres trigram/cosine/levenshtein distance
        qs = super().get_queryset()
        items = qs.values_list("id", "name", "code")
        res = []
        for item in items:
            obj = {}
            similar_items = []
            for id, name, code in items:
                if fuzz.ratio(item[1], name) > 80:
                    similar_items.append(
                        {
                            "id": id,
                            "name": f"{name} ({code})",
                            "code": code,
                            # "match": fuzz.ratio(item[1], name)
                        }
                    )
            if len(similar_items) > 1:
                obj["items"] = similar_items
                obj["config"] = {}
                # res[item[1]] = sim
                res.append(obj)

        unique_ids = {}

        # Filter out duplicate items
        filtered_data = []
        index = 0
        for group in res:
            unique_items = []
            for item in group["items"]:
                item_id = item["id"]
                if item_id not in unique_ids:
                    unique_ids[item_id] = True
                    unique_items.append(item)
            if unique_items:
                index += 1
                filtered_data.append(
                    {
                        "items": [x["id"] for x in unique_items],
                        "config": group["config"],
                        "index": index,
                    }
                )

        return Response(filtered_data)

    @action(detail=False, methods=["POST"])
    def merge(self, request, *args, **kwargs):
        groups_not_merged = []
        items_not_merged = []
        flag = False
        for index, group in enumerate(request.data):
            if group.get("config"):
                ret = self.merge_items(group["items"], group["config"])
            else:
                ret = self.merge_items(group["items"])
            if ret:
                flag = True
                groups_not_merged.append(index + 1)
                items_not_merged.append(group)
        if flag:
            res = {
                "error": {
                    "message": f"Items in Groups {','.join([str(x) for x in groups_not_merged])} were not merged due to conflicting config on items.",
                    "items": items_not_merged,
                }
            }
            return Response(res, status=209)
        return Response(status=200)

    @action(detail=True)
    def details(self, request, pk=None, *args, **kwargs):
        qs = (
            super()
            .get_queryset()
            .select_related(
                "account",
                "sales_account",
                "purchase_account",
                "discount_allowed_account",
                "discount_received_account",
                "expense_account",
                "fixed_asset_account",
                "tax_scheme",
            )
        )
        item = get_object_or_404(queryset=qs, pk=pk)
        serializer = ItemDetailSerializer(item, context={"request": request}).data
        return Response(serializer)

    # items listing for POS
    @action(detail=False)
    def pos(self, request, *args, **kwargs):
        self.filter_backends = (rf_filters.SearchFilter,)
        self.queryset = self.get_queryset().filter(can_be_sold=True)
        self.serializer_class = ItemPOSSerializer
        self.search_fields = [
            "name",
            "code",
            "description",
            "search_data",
        ]
        self.paginator.page_size = settings.POS_ITEMS_SIZE
        return super().list(request)

    @action(detail=False, url_path="sales-choices")
    def sales_choices(self, request, *args, **kwargs):
        queryset = self.get_queryset().filter(can_be_sold=True)
        serializer = GenericSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, url_path="purchase-choices")
    def purchase_choices(self, request, *args, **kwargs):
        queryset = self.get_queryset().filter(can_be_purchased=True)
        serializer = GenericSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["POST"], url_path="import")
    def import_from_file(self, request, *args, **kwargs):
        file = request.FILES["file"]
        wb = load_workbook(file)
        ws = wb.active
        ws.delete_rows(1)
        items = []

        def get_bool_for_txt(value):
            if not value:
                return False
            elif value.upper() == "F":
                return False
            elif value.upper() == "T":
                return True
            else:
                raise ValueError("Invalid value detected.")

        def get_category(value):
            if not value:
                return None
            else:
                qs = Category.objects.filter(
                    name__iexact=value, company_id=request.company.id
                )
                if qs.exists():
                    return qs.first()
                else:
                    if request.data.get("create_new_category") == "true":
                        cat = Category.objects.create(
                            name=value, company_id=request.company.id
                        )
                        return cat
                    return None

        def get_Tax(value):
            if not value:
                return None
            else:
                qs = TaxScheme.objects.filter(
                    name__iexact=value, company_id=request.company.id
                )
                if qs.exists():
                    return qs.first()
                else:
                    return None

        def get_unit(value):
            if not value:
                return None
            else:
                qs = Unit.objects.filter(
                    name__iexact=value, company_id=request.company.id
                )
                if qs.exists():
                    return qs.first()
                else:
                    if request.data.get("create_new_unit") == "true":
                        new_unit = Unit.objects.create(
                            name=value, company_id=request.company.id
                        )
                        return new_unit
                    return None

        for row in ws.iter_rows(values_only=True):
            item = {
                "company_id": request.company.id,
                "name": row[0],
                "code": row[1],
                "category": get_category(row[2]),
                "cost_price": row[3],
                "selling_price": row[4],
                "can_be_purchased": get_bool_for_txt(row[5]),
                "can_be_sold": get_bool_for_txt(row[6]),
                "track_inventory": get_bool_for_txt(row[7]),
                "tax_scheme": get_Tax(row[8]),
                "unit": get_unit(row[9]),
            }
            item_obj = Item(**item)
            items.append(item_obj)

        with transaction.atomic():
            try:
                Item.objects.bulk_create(items, batch_size=1000)
                # items = Item.objects.bulk_create(items)
                accounts_to_create = []
                items_to_update = []
                for i, item in enumerate(items):
                    # print(i)
                    if item.can_be_purchased:
                        name = item.name + " (Purchase)"
                        purchase_account = Account(name=name, company=item.company)
                        purchase_account.add_category(acc_cat_system_codes["Purchase"])
                        purchase_account.suggest_code(item)
                        accounts_to_create.append(purchase_account)
                        item.purchase_account = purchase_account
                        item.purchase_account_type = "dedicated"
                        item.dedicated_purchase_account = purchase_account

                        name = "Discount Received - " + item.name
                        discount_received_acc = Account(name=name, company=item.company)
                        discount_received_acc.add_category(
                            acc_cat_system_codes["Discount Income"]
                        )
                        discount_received_acc.suggest_code(item)
                        accounts_to_create.append(discount_received_acc)
                        item.discount_received_account = discount_received_acc
                        item.discount_received_account_type = "dedicated"
                        item.dedicated_discount_received_account = discount_received_acc

                    if item.can_be_sold:
                        name = item.name + " (Sales)"
                        sales_account = Account(name=name, company=item.company)
                        sales_account.add_category(acc_cat_system_codes["Sales"])
                        sales_account.suggest_code(item)
                        accounts_to_create.append(sales_account)
                        item.sales_account = sales_account
                        item.sales_account_type = "dedicated"
                        item.dedicated_sales_account = sales_account

                        name = "Discount Allowed - " + item.name
                        discount_allowed_account = Account(
                            name=name, company=item.company
                        )
                        discount_allowed_account.add_category(
                            acc_cat_system_codes["Discount Expenses"]
                        )
                        discount_allowed_account.suggest_code(item)
                        accounts_to_create.append(discount_allowed_account)
                        item.discount_allowed_account = discount_allowed_account
                        item.discount_allowed_account_type = "dedicated"
                        item.dedicated_discount_allowed_account = (
                            discount_allowed_account
                        )

                    items_to_update.append(item)

                Account.objects.bulk_create(accounts_to_create, batch_size=1000)
                Item.objects.bulk_update(
                    items_to_update,
                    fields=[
                        "purchase_account",
                        "sales_account",
                        "discount_received_account",
                        "discount_allowed_account",
                        "purchase_account_type",
                        "dedicated_purchase_account",
                        "discount_received_account_type",
                        "dedicated_discount_received_account",
                        "sales_account_type",
                        "dedicated_sales_account",
                        "discount_allowed_account_type",
                        "dedicated_discount_allowed_account",
                    ],
                )

            except IntegrityError as e:
                code = e.args[0].split("=(")[1].split(")")[0].split(",")[0]
                res_msg = f"Duplicate items with code {code} detected."
                return Response({"details": res_msg}, status=400)
        return Response({}, status=200)

    @action(detail=True, methods=["GET"], url_path="available-stock")
    def available_stock_data(self, request, pk=None, *args, **kwargs):
        item = self.get_object()
        return Response(item.available_stock_data, status=200)

    def export_stock_movement_report(self, serializer):
        wb = Workbook("stock_movement.xlsx")
        ws = wb.add_worksheet("Stock Movement")

        # Define formats
        header_format = wb.add_format({"bold": True, "align": "center", "border": 1})
        cell_format = wb.add_format({"align": "center", "border": 1})
        currency_format = wb.add_format({"num_format": "#,##0.00", "border": 1})

        # Define table headers
        headers = [
            "ITEM DESCRIPTION",
            "OPENING BALANCE",
            "PURCHASE",
            "PURCHASE RETURN",
            "SALES",
            "SALES RETURN",
            "STOCK IN",
            "STOCK OUT",
            "PRODUCTION",
            "CONSUMPTION",
            "CLOSING STOCK",
        ]
        sub_headers = [
            ["ITEM CODE", "ITEM NAME", "UNIT"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
            ["QTY", "VALUE"],
        ]

        # Write main headers
        col = 0
        for idx, header in enumerate(headers):
            col_span = len(sub_headers[idx])
            ws.merge_range(0, col, 0, col + col_span - 1, header, header_format)
            col += col_span

        # Write sub-headers
        col = 0
        for sub_header in sub_headers:
            for sub in sub_header:
                ws.write(1, col, sub, header_format)
                col += 1

        # Write data rows
        row = 2
        for row_data in serializer.data:
            col = 0
            for key, value in row_data.items():
                if key in ["id", "account"]:
                    print(key, value)
                    continue
                if isinstance(value, (int, float)):
                    ws.write(
                        row, col, value, currency_format if col > 2 else cell_format
                    )
                else:
                    ws.write(row, col, value, cell_format)
                col += 1
            row += 1

        # Save workbook
        wb.close()

        # Serve the file
        with open("stock_movement.xlsx", "rb") as f:
            data = f.read()

        response = HttpResponse(
            data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=stock_movement.xlsx"
        return response

    @action(detail=False, methods=["get"], url_path="stock-movement")
    def stock_movement(self, request, pk=None, *args, **kwargs):
        start_date = request.query_params.get("start_date")
        if not start_date:
            start_date = request.company.current_fiscal_year.start_date.strftime(
                "%Y-%m-%d"
            )

        end_date = request.query_params.get("end_date")
        if not end_date:
            end_date = datetime.now().date().strftime("%Y-%m-%d")

        item_ids = request.query_params.get("item_ids", None)
        if item_ids:
            item_ids = item_ids.split(",")

        export = request.query_params.get("export", "false") == "true"
        export_type: Literal["xlsx", None] = request.query_params.get("type", None)

        # validate start_date and end_date
        try:
            if start_date:
                start_date = datetime.strptime(start_date, "%Y-%m-%d")
            if end_date:
                end_date = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise ValidationError(
                "Invalid date format. Date format should be YYYY-MM-DD."
            )
        else:
            # validate start_date is not greater than end_date
            if start_date and start_date > end_date:
                raise ValidationError("Start date cannot be greater than end date.")

        base_qs = (
            Item.objects.filter(
                company=request.company,
                track_inventory=True,
            )
            .values("id", "account_id")
            .order_by("id")
        )

        if item_ids:
            base_qs = base_qs.filter(id__in=item_ids)

        paginated = (
            base_qs
            if export and export_type == "xlsx"
            else self.paginate_queryset(base_qs)
        )

        # fifo_enabled = request.company.inventory_setting.enable_fifo

        txn_raw_query = f"""
            WITH RECURSIVE weight_calc AS (
                SELECT
                    ROW_NUMBER() OVER (PARTITION BY account_id ORDER BY product_journalentry.date, product_transaction.id) AS rn,
                    product_transaction.id AS id,
                    dr_amount,
                    cr_amount,
                    account_id,
                    django_content_type.model AS content_type,
                    CASE
                        WHEN dr_amount IS NOT NULL THEN rate
                    END AS entered_rate,
                    COALESCE(dr_amount, cr_amount * -1) AS weight,
                    CASE
                        WHEN product_journalentry.date < {'%s'} THEN 'opening'
                        ELSE 'closing'
                    END AS period
                FROM product_transaction
                JOIN product_journalentry
                    ON product_transaction.journal_entry_id = product_journalentry.id
                JOIN django_content_type
                    ON product_journalentry.content_type_id = django_content_type.id
                WHERE
                    account_id IN ({', '.join(['%s'] * len(paginated))})
                    AND product_journalentry.date <= {'%s'}
            ),
            running_calcs AS (
                SELECT
                    *,
                    SUM(weight) OVER (PARTITION BY account_id ORDER BY rn) AS current_balance
                FROM weight_calc
            ),
            final_calc AS (
                SELECT
                    rn,
                    id,
                    account_id,
                    dr_amount,
                    cr_amount,
                    entered_rate,
                    weight,
                    current_balance,
                CASE
                    WHEN content_type = 'debitnoterow' THEN entered_rate
                    WHEN content_type = 'creditnoterow' THEN 0
                    WHEN dr_amount IS NOT NULL THEN entered_rate
                    ELSE 0
                END AS supposed_rate,
                CASE
                    WHEN content_type = 'debitnoterow' THEN entered_rate
                    WHEN content_type = 'creditnoterow' THEN 0
                    WHEN dr_amount IS NOT NULL THEN entered_rate
                    ELSE 0
                END AS calculated_rate,
                period
                FROM running_calcs
                WHERE rn = 1

                UNION ALL

                SELECT
                    rc.rn,
                    rc.id,
                    rc.account_id,
                    rc.dr_amount,
                    rc.cr_amount,
                    rc.entered_rate,
                    rc.weight,
                    rc.current_balance,
                CASE
                    WHEN content_type = 'debitnoterow' THEN rc.entered_rate
                    WHEN content_type = 'creditnoterow' THEN fc.calculated_rate
                    WHEN rc.dr_amount IS NOT NULL THEN rc.entered_rate
                    ELSE fc.calculated_rate
                END AS supposed_rate,
                CASE
                    WHEN rc.current_balance <> 0 THEN
                        (
                            (fc.calculated_rate * fc.current_balance) +
                            (
                                rc.weight *
                                CASE
                                    WHEN content_type = 'debitnoterow' THEN rc.entered_rate
                                    WHEN content_type = 'creditnoterow' THEN fc.calculated_rate
                                    WHEN rc.dr_amount IS NOT NULL THEN rc.entered_rate
                                    ELSE fc.calculated_rate
                                END
                            )
                        ) / rc.current_balance
                    ELSE 0
                END AS calculated_rate,
                rc.period
                FROM running_calcs rc
                JOIN final_calc fc
                    ON rc.rn = fc.rn + 1 AND rc.account_id = fc.account_id
            ),
            ranked_rows AS (
                SELECT
                    *,
                    ROW_NUMBER() OVER (PARTITION BY account_id, period ORDER BY rn DESC) AS row_num
                FROM final_calc
            )
            SELECT
                rn,
                account_id,
                id,
                dr_amount,
                cr_amount,
                ROUND(COALESCE(current_balance::numeric, 0), 4) AS current_balance,
                ROUND(COALESCE(calculated_rate::numeric, 0), 4) AS calculated_rate,
                period
            FROM ranked_rows
            WHERE row_num = 1
            ORDER BY account_id, period DESC;
        """

        item_ids = [x["id"] for x in paginated]
        account_ids = [x["account_id"] for x in paginated]

        opening_closing = Transaction.objects.raw(
            txn_raw_query,
            [start_date.strftime("%Y-%m-%d")]
            + account_ids
            + [end_date.strftime("%Y-%m-%d")],
        )

        opening_closing_dict = defaultdict(
            lambda: {
                "opening_rate": 0,
                "closing_rate": 0,
                "opening_qty": 0,
                "closing_qty": 0,
                "opening_value": 0,
                "closing_value": 0,
            }
        )

        for oc in opening_closing:
            if oc.period == "opening":
                # assume that opening is closing as well, for cases where there is no closing row for an account
                opening_closing_dict[oc.account_id][
                    "closing_rate"
                ] = opening_closing_dict[oc.account_id]["opening_rate"] = (
                    oc.calculated_rate
                )
                opening_closing_dict[oc.account_id][
                    "closing_qty"
                ] = opening_closing_dict[oc.account_id]["opening_qty"] = (
                    oc.current_balance
                )
                opening_closing_dict[oc.account_id][
                    "closing_value"
                ] = opening_closing_dict[oc.account_id]["opening_value"] = (
                    oc.current_balance * oc.calculated_rate
                )
            elif oc.period == "closing":
                opening_closing_dict[oc.account_id]["closing_rate"] = oc.calculated_rate
                opening_closing_dict[oc.account_id]["closing_qty"] = oc.current_balance
                opening_closing_dict[oc.account_id]["closing_value"] = (
                    oc.current_balance * oc.calculated_rate
                )

        item_raw_query = f"""
        SELECT
            "product_item"."id" as "id",
            "product_item"."name",
            "product_item"."code",
            "product_item"."unit_id",
            "product_item"."account_id",
            COALESCE((
                SELECT SUM(U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'purchasevoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "purchase_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'purchasevoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "purchase_value",
            COALESCE((
                SELECT SUM(U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'debitnoterow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "purchase_return_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'debitnoterow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "purchase_return_value",
            COALESCE((
                SELECT SUM(U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'salesvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "sales_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'salesvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "sales_value",
            COALESCE((
                SELECT SUM(U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'creditnoterow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "sales_return_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'creditnoterow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "sales_return_value",
            COALESCE((
                SELECT SUM(U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryadjustmentvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "stock_in_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryadjustmentvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "stock_in_value",
            COALESCE((
                SELECT SUM(U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryadjustmentvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "stock_out_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryadjustmentvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "stock_out_value",
            COALESCE((
                SELECT SUM(U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryconversionvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "production_qty",
            COALESCE((
                SELECT SUM(U0."rate" * U0."dr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryconversionvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "production_value",
            COALESCE((
                SELECT SUM(U0."cr_amount")
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id"
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryconversionvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "consumption_qty",
            COALESCE((
                SELECT SUM(
                    (
                        CASE
                            WHEN jsonb_array_length(value) >= 2 THEN
                                -- Calculate weighted average: quantity * rate / quantity = rate
                                (value->0)::numeric * (value->1)::numeric / NULLIF((value->0)::numeric, 0)
                            ELSE 0
                        END
                    ) * U0.cr_amount
                )
                FROM "product_transaction" U0
                INNER JOIN "product_journalentry" U2 ON U0."journal_entry_id" = U2."id"
                INNER JOIN "django_content_type" U3 ON U2."content_type_id" = U3."id",
                jsonb_each(
                    CASE
                        WHEN U0.consumption_data = '{{}}'::jsonb THEN '{{"0":[0,0]}}'::jsonb
                        WHEN U0.consumption_data IS NULL THEN '{{"0":[0,0]}}'::jsonb
                        ELSE U0.consumption_data
                    END
                ) AS j(key, value)
                WHERE
                    U0."account_id" = "product_item"."account_id"
                    AND U3."model" = 'inventoryconversionvoucherrow'
                    AND U2."date" >= {'%s'}::date
                    AND U2."date" <= {'%s'}::date
            ), 0) AS "consumption_value",
            "product_unit"."id",
            "product_unit"."name",
            "product_unit"."short_name"
        FROM "product_item"
        LEFT OUTER JOIN "product_unit"
            ON "product_item"."unit_id" = "product_unit"."id"
        WHERE
            "product_item"."id" IN ({', '.join(['%s'] * len(paginated))})
            AND "product_item"."track_inventory";
        """

        qs = Item.objects.raw(
            item_raw_query,
            (
                [
                    start_date.strftime("%Y-%m-%d"),
                    end_date.strftime("%Y-%m-%d"),
                ]
                * 16
            )
            + item_ids,
        )

        for item in qs:
            setattr(
                item,
                "opening_rate",
                opening_closing_dict[item.account_id]["opening_rate"],
            )
            setattr(
                item,
                "opening_qty",
                opening_closing_dict[item.account_id]["opening_qty"],
            )
            setattr(
                item,
                "opening_value",
                opening_closing_dict[item.account_id]["opening_value"],
            )
            setattr(
                item,
                "closing_rate",
                opening_closing_dict[item.account_id]["closing_rate"],
            )
            setattr(
                item,
                "closing_qty",
                opening_closing_dict[item.account_id]["closing_qty"],
            )
            setattr(
                item,
                "closing_value",
                opening_closing_dict[item.account_id]["closing_value"],
            )

        serializer = StockMovementSerializer(qs, many=True)

        if export and export_type == "xlsx":
            return self.export_stock_movement_report(serializer)

        return self.get_paginated_response(serializer.data)


class ItemOpeningBalanceViewSet(DestroyModelMixin, CRULViewSet):
    serializer_class = ItemOpeningSerializer
    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = [
        "item__name",
        "item__code",
        "item__description",
        "item__search_data",
        "opening_balance",
    ]
    filterset_class = InventoryAccountFilterSet

    collections = (
        (
            "items",
            Item.objects.only("id", "name").filter(
                track_inventory=True, account__opening_balance=0
            ),
            GenericSerializer,
            True,
            ["name"],
        ),
    )

    def get_queryset(self, company_id=None):
        return super().get_queryset().filter(opening_balance__gt=0)

    def get_update_defaults(self, request=None, *args, **kwargs):
        self.collections = (
            (
                "items",
                Item.objects.filter(track_inventory=True).filter(
                    Q(account__opening_balance=0) | Q(account_id=self.kwargs.get("pk"))
                ),
                ItemListSerializer,
                True,
                ["name"],
            ),
        )
        return self.get_defaults(request=request)

    def create(self, request, *args, **kwargs):
        data = request.data
        account = get_object_or_404(
            InventoryAccount,
            item__id=data.get("item_id"),
            opening_balance=0,
            company=request.company,
        )
        opening_balance = data.get("opening_balance", None)
        if not opening_balance:
            raise ValidationError({"opening_balance": ["Opening balance is required."]})
        account.opening_balance = data.get("opening_balance")
        account.opening_balance_rate = data.get("opening_balance_rate")
        account.save()
        fiscal_year = self.request.company.current_fiscal_year
        account.item.update_opening_balance(fiscal_year)
        return Response({})

    def perform_update(self, serializer):
        super().perform_update(serializer)
        fiscal_year = self.request.company.current_fiscal_year
        serializer.instance.item.update_opening_balance(fiscal_year)

    def destroy(self, request, *args, **kwargs):
        account = self.get_object()
        account.opening_balance = 0
        account.opening_balance_rate = 0
        account.save()
        fiscal_year = self.request.company.current_fiscal_year
        account.item.update_opening_balance(fiscal_year)
        JournalEntry.objects.filter(
            content_type__model="item",
            content_type__app_label="product",
            object_id=account.item.id,
        ).delete()
        return Response({})


class BookViewSet(InputChoiceMixin, CRULViewSet):
    collections = (("brands", Brand, BrandSerializer, True, ["name"]),)

    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = [
        "name",
        "code",
        "description",
        "search_data",
        "selling_price",
        "cost_price",
    ]
    filterset_class = BookFilterSet

    def get_queryset(self, **kwargs):
        queryset = Item.objects.filter(
            category__name="Book", company=self.request.company
        )
        return queryset

    serializer_class = BookSerializer

    @action(detail=False)
    def category(self, request, *args, **kwargs):
        cat = Category.objects.filter(company=self.request.company, name="Book").first()
        return Response(InventoryCategorySerializer(cat).data)


class UnitViewSet(InputChoiceMixin, ShortNameChoiceMixin, CRULViewSet):
    serializer_class = UnitSerializer


class InventoryCategoryViewSet(InputChoiceMixin, ShortNameChoiceMixin, CRULViewSet):
    serializer_class = InventoryCategorySerializer
    collections = (
        ("units", Unit, UnitSerializer, True, ["name", "short_name"]),
        ("accounts", Account, AccountMinSerializer, True, ["name"]),
        # ('purchase_accounts', Account.objects.filter(category__name="Purchase"), AccountMinSerializer),
        # ('sales_accounts', Account.objects.filter(category__name="Sales"), AccountMinSerializer),
        ("tax_scheme", TaxScheme, TaxSchemeMinSerializer, True, ["name"]),
        (
            "discount_allowed_accounts",
            Account.objects.filter(
                category__system_code=acc_cat_system_codes["Discount Expenses"]
            ),
            AccountMinSerializer,
            True,
            ["name"],
        ),
        (
            "discount_received_accounts",
            Account.objects.filter(
                category__system_code=acc_cat_system_codes["Discount Income"]
            ),
            AccountMinSerializer,
            True,
            ["name"],
        ),
    )

    def get_defaults(self, request=None):
        acc_system_codes = settings.ACCOUNT_SYSTEM_CODES
        system_codes = [
            acc_system_codes["Sales Account"],
            acc_system_codes["Purchase Account"],
            acc_system_codes["Discount Expenses"],
            acc_system_codes["Discount Income"],
        ]

        accounts = Account.objects.filter(
            company=request.company, system_code__in=system_codes
        ).values_list("system_code", "id")
        account_ids = {system_code: id for system_code, id in accounts}
        return {
            "options": {
                "global_accounts": {
                    "sales_account_id": account_ids.get(
                        acc_system_codes["Sales Account"]
                    ),
                    "purchase_account_id": account_ids.get(
                        acc_system_codes["Purchase Account"]
                    ),
                    "discount_allowed_account_id": account_ids.get(
                        acc_system_codes["Discount Expenses"]
                    ),
                    "discount_received_account_id": account_ids.get(
                        acc_system_codes["Discount Income"]
                    ),
                }
            },
        }

    @transaction.atomic
    def perform_update(self, request, *args, **kwargs):
        category = self.get_object()
        apply_changes = self.request.query_params.get("Account Changes Detected")
        if (
            (
                category.items_sales_account_type
                != self.request.data.get("items_sales_account_type")
            )
            or (
                category.items_purchase_account_type
                != self.request.data.get("items_purchase_account_type")
            )
            or (
                category.items_discount_allowed_account_type
                != self.request.data.get("items_discount_allowed_account_type")
            )
            or (
                category.items_discount_received_account_type
                != self.request.data.get("items_discount_received_account_type")
            )
        ) and apply_changes is None:
            raise UnprocessableException(
                detail="Changes in account settings detected! Do you want the changes to be applied to the items belonging to this category?",
                code="Account Changes Detected",
            )
        super().perform_update(request, *args, **kwargs)
        if apply_changes == "true":
            category.refresh_from_db()
            category.apply_account_settings_to_items()

    def get_collections(self, request=None, *args, **kwargs):
        collections_data = super().get_collections(self.request)
        collections_data["fixed_assets_categories"] = GenericSerializer(
            AccountCategory.objects.get(
                system_code=acc_cat_system_codes["Fixed Assets"],
                company=self.request.company,
            ).get_descendants(include_self=True),
            many=True,
        ).data
        collections_data["direct_expenses_categories"] = GenericSerializer(
            AccountCategory.objects.get(
                system_code=acc_cat_system_codes["Direct Expenses"],
                company=self.request.company,
            ).get_descendants(include_self=True),
            many=True,
        ).data
        collections_data["indirect_expenses_categories"] = GenericSerializer(
            AccountCategory.objects.get(
                system_code=acc_cat_system_codes["Indirect Expenses"],
                company=self.request.company,
            ).get_descendants(include_self=True),
            many=True,
        ).data
        return collections_data

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action == "list":
            qs = qs.order_by("-id")
        return qs

    def get_serializer_class(self):
        if self.action == "retrieve":
            return InventoryCategoryFormSerializer
        return super().get_serializer_class()

    @action(detail=False, url_path="trial-balance")
    def trial_balance(self, request, *args, **kwargs):
        qs = self.get_queryset().filter(Q(track_inventory=True) | Q(fixed_asset=True))
        return Response(InventoryCategoryTrialBalanceSerializer(qs, many=True).data)


class BrandViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = BrandSerializer


class InventoryAccountViewSet(InputChoiceMixin, CRULViewSet):
    serializer_class = InventoryAccountSerializer

    filter_backends = (
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    )
    search_fields = (
        "code",
        "name",
    )

    def get_account_ids(self, obj):
        return [obj.id]

    @action(detail=True, methods=["get"], url_path="journal-entries")
    def journal_entries(self, request, pk=None, *args, **kwargs):
        param = request.GET
        start_date = param.get("start_date")
        end_date = param.get("end_date")
        obj = self.get_object()
        entries = (
            JournalEntry.objects.filter(transactions__account_id=obj.pk)
            .order_by("pk", "date")
            .prefetch_related("transactions", "content_type", "transactions__account")
            .select_related()
        )

        if start_date or end_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.strptime(end_date, "%Y-%m-%d")

            if start_date == end_date:
                entries = entries.filter(date=start_date)
            else:
                entries = entries.filter(date__range=[start_date, end_date])
        serializer = JournalEntrySerializer(
            entries, context={"account": obj}, many=True
        )
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def transactions(self, request, pk=None, *args, **kwargs):
        param = request.GET
        obj = self.get_object()
        serializer_class = self.get_serializer_class()
        data = serializer_class(obj).data
        account_ids = self.get_account_ids(obj)
        start_date = param.get("start_date", None)
        end_date = param.get("end_date", None)
        transactions = (
            Transaction.objects.filter(account_id__in=account_ids)
            .order_by("-journal_entry__date", "-pk")
            .select_related("journal_entry__content_type")
        )

        aggregate = {}
        if start_date or end_date:
            if start_date == "null":
                from rest_framework.exceptions import ValidationError

                raise ValidationError(
                    "Either start date or both dates are required to filter."
                )
            start_date = datetime.strptime(start_date, "%Y-%m-%d")
            # TODO: if only start date is given, raise error or process some other way?
            end_date = (
                datetime.strptime(end_date, "%Y-%m-%d")
                if end_date != "null"
                else datetime.today()
            )
            if start_date == end_date:
                transactions = transactions.filter(journal_entry__date=start_date)
            else:
                transactions = transactions.filter(
                    journal_entry__date__range=[start_date, end_date]
                )
            aggregate = transactions.aggregate(Sum("dr_amount"), Sum("cr_amount"))

        # Only show 5 because fetching voucher_no is expensive because of GFK, GFK to be cached
        # self.paginator.page_size = 5
        page = self.paginate_queryset(transactions)
        serializer = TransactionEntrySerializer(page, many=True)
        data["transactions"] = self.paginator.get_response_data(serializer.data)
        data["aggregate"] = aggregate
        return Response(data)

    @action(detail=False, url_path="trial-balance")
    def trial_balance(self, request, format=None, *args, **kwargs):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        if start_date and end_date:
            qq = (
                InventoryAccount.objects.filter(company=request.company)
                .annotate(
                    od=Sum(
                        "transactions__dr_amount",
                        filter=Q(transactions__journal_entry__date__lt=start_date),
                    ),
                    oc=Sum(
                        "transactions__cr_amount",
                        filter=Q(transactions__journal_entry__date__lt=start_date),
                    ),
                    cd=Sum(
                        "transactions__dr_amount",
                        filter=Q(transactions__journal_entry__date__lte=end_date),
                    ),
                    cc=Sum(
                        "transactions__cr_amount",
                        filter=Q(transactions__journal_entry__date__lte=end_date),
                    ),
                )
                .values("id", "name", "item__category_id", "od", "oc", "cd", "cc")
                .exclude(od=None, oc=None, cd=None, cc=None)
            )
            return Response(list(qq))
        return Response({})


class InventorySettingsViewSet(CRULViewSet):
    serializer_class = InventorySettingCreateSerializer

    def get_defaults(self, request=None, *args, **kwargs):
        i_setting = self.request.company.inventory_setting
        data = {"fields": self.get_serializer(i_setting).data}
        return data


class BillOfMaterialViewSet(DeleteRows, CRULViewSet):
    model = BillOfMaterial
    row = BillOfMaterialRow
    serializer_class = BillOfMaterialCreateSerializer
    collections = [
        [
            "finished_products",
            Item.objects.only("id", "name").filter(
                track_inventory=True, bill_of_material__isnull=True
            ),
            GenericSerializer,
            True,
            ["name"],
        ],
        [
            "units",
            Unit,
            GenericSerializer,
            True,
            ["name"],
        ],
        [
            "items",
            Item.objects.only("id", "name").filter(track_inventory=True),
            GenericSerializer,
            True,
            ["name"],
        ],
    ]
    filter_backends = [
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "finished_product__name",
    ]

    def get_queryset(self, **kwargs):
        qs = super(BillOfMaterialViewSet, self).get_queryset()
        return qs.order_by(
            "-created_at",
        )

    def get_serializer_class(self):
        if self.action == "list":
            return BillOfMaterialListSerializer
        return BillOfMaterialCreateSerializer


class InventoryAdjustmentVoucherViewSet(DeleteRows, CRULViewSet):
    queryset = InventoryAdjustmentVoucher.objects.all()
    serializer_class = InventoryAdjustmentVoucherCreateSerializer
    model = InventoryAdjustmentVoucher
    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    filterset_class = InventoryAdjustmentVoucherFilterSet
    search_fields = [
        "remarks",
        "total_amount",
        "date",
        "purpose",
        "voucher_no",
    ]

    def get_queryset(self, **kwargs):
        qs = super(InventoryAdjustmentVoucherViewSet, self).get_queryset()
        if self.action == "retrieve":
            qs = qs.prefetch_related("rows__item", "rows__unit")
        return qs.order_by("-date", "-voucher_no")

    def get_serializer_class(self):
        if self.action == "list":
            return InventoryAdjustmentVoucherListSerializer
        elif self.action == "retrieve":
            return InventoryAdjustmentVoucherDetailSerializer
        return InventoryAdjustmentVoucherCreateSerializer

    collections = [
        (
            "items",
            Item.objects.only("id", "name").filter(track_inventory=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "units",
            Unit,
            GenericSerializer,
            True,
            ["name"],
        ),
    ]

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk, *args, **kwargs):
        inventory_adjustment_voucher = self.get_object()
        message = request.data.get("message")
        if not message:
            raise ValidationError(
                {"message": "message field is required for cancelling voucher!"}
            )
        inventory_adjustment_voucher.cancel(message=message)
        return Response({})

    @action(detail=False, methods=["POST"], url_path="import")
    def import_from_file(self, request, *args, **kwargs):
        file = request.FILES["file"]
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        ws.delete_rows(1)
        items = []
        unadjusted_items = []

        def get_item(name):
            if not name:
                return None
            else:
                qs = Item.objects.filter(
                    code__iexact=str(name), company_id=request.company.id
                ).select_related("unit")
                if qs.exists():
                    return qs.first()
                else:
                    return None

        for row in ws.iter_rows(values_only=True):
            if not get_item(row[0]):
                unadjusted_items.append(row[0])
            else:
                item = {
                    "company_id": request.company.id,
                    "item_id": get_item(row[0]).id,
                    "unit_id": get_item(row[0]).unit.id
                    if get_item(row[0]).unit
                    else None,
                    "quantity": row[1],
                    "rate": row[2],
                    "description": row[3],
                    "selected_item_obj": (GenericSerializer(get_item(row[0]))).data,
                    "selected_unit_obj": (GenericSerializer(get_item(row[0]).unit)).data
                    if get_item(row[0]).unit
                    else None,
                }
                items.append(item)
        response = {"items": items, "unadjusted_items": unadjusted_items}
        return Response(response, status=200)

    @action(detail=True, url_path="journal-entries")
    def journal_entries(self, request, pk, *args, **kwargs):
        inventory_adjustment_voucher = get_object_or_404(
            InventoryAdjustmentVoucher, pk=pk
        )
        journals = inventory_adjustment_voucher.journal_entries()
        return Response(SalesJournalEntrySerializer(journals, many=True).data)


class InventoryConversionVoucherViewSet(DeleteRows, CRULViewSet):
    queryset = InventoryConversionVoucher.objects.all()
    serializer_class = InventoryConversionVoucherCreateSerializer
    model = InventoryConversionVoucher
    filter_backends = [
        filters.DjangoFilterBackend,
        rf_filters.OrderingFilter,
        rf_filters.SearchFilter,
    ]
    search_fields = [
        "voucher_no",
        "date",
        "finished_product__finished_product__name",
        "rows__item__name",
    ]
    filterset_class = InventoryConversionVoucherFilterSet

    def get_queryset(self, **kwargs):
        qs = super(InventoryConversionVoucherViewSet, self).get_queryset()
        return qs.order_by("-date", "-voucher_no")

    def get_serializer_class(self):
        if self.action == "list":
            return InventoryConversionVoucherListSerializer
        elif self.action == "retrieve":
            return InventoryConversionVoucherDetailSerializer
        return InventoryConversionVoucherCreateSerializer

    @action(detail=True, methods=["POST"])
    def cancel(self, request, pk, *args, **kwargs):
        inventory_conversion_voucher = self.get_object()
        message = request.data.get("message")
        if not message:
            raise ValidationError(
                {"message": "message field is required for cancelling voucher!"}
            )
        inventory_conversion_voucher.cancel(message=message)
        return Response({})

    collections = [
        (
            "items",
            Item.objects.only("id", "name").filter(track_inventory=True),
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "units",
            Unit,
            GenericSerializer,
            True,
            ["name"],
        ),
        (
            "finished_products",
            BillOfMaterial.objects.prefetch_related("finished_product").only(
                "id", "finished_product"
            ),
            GenericSerializer,
            True,
            ["name"],
        ),
    ]
